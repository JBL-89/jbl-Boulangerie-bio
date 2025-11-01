#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Application SARL AQUI BIO PAIN - Version Streamlit
Portage de l'application Flask vers Streamlit
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
import sqlite3
import hashlib
import os
import base64
from pathlib import Path
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import xlsxwriter
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Configuration de la page
st.set_page_config(
    page_title="SARL AQUI BIO PAIN",
    page_icon="🥖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# === CONFIGURATION DATABASE ===
# Configuration pour différents environnements
DATABASE_PATH = os.getenv('DATABASE_PATH', 'boulangerie_streamlit.db')

# Configuration pour la production avec Streamlit Cloud
try:
    if hasattr(st, 'secrets') and 'database' in st.secrets:
        if 'path' in st.secrets['database']:
            DATABASE_PATH = st.secrets['database']['path']
except Exception:
    pass  # Utiliser la configuration par défaut

# Mode production
PRODUCTION_MODE = os.getenv('STREAMLIT_ENV') == 'production'

# === UTILITAIRES ===
def get_base64_of_bin_file(bin_file):
    """Convertit un fichier en base64 pour l'affichage"""
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def load_css():
    """Charge le CSS personnalisé pour reproduire le style de l'application Flask"""
    
    # Vérification et chargement de l'image de fond
    pain_image_path = Path("static/images/pain1.avif")
    background_image = ""
    
    if pain_image_path.exists():
        try:
            bin_str = get_base64_of_bin_file(pain_image_path)
            background_image = f"""
            background: linear-gradient(rgba(255, 248, 220, 0.7), rgba(255, 248, 220, 0.7)),
                        url("data:image/avif;base64,{bin_str}");
            background-size: cover;
            background-attachment: fixed;
            background-repeat: no-repeat;
            background-position: center;
            """
        except Exception as e:
            st.warning(f"Impossible de charger l'image de fond: {e}")
            background_image = "background: linear-gradient(rgba(255, 248, 220, 0.9), rgba(255, 248, 220, 0.9));"
    else:
        background_image = "background: linear-gradient(rgba(255, 248, 220, 0.9), rgba(255, 248, 220, 0.9));"
    
    css = f"""
    <style>
    /* Import des styles personnalisés de l'application boulangerie */
    :root {{
        --primary-color: #8B4513;
        --secondary-color: #D2691E;
        --accent-color: #F4A460;
        --success-color: #198754;
        --warning-color: #ffc107;
        --danger-color: #dc3545;
        --info-color: #0dcaf0;
        --dark-color: #212529;
        --light-color: #f8f9fa;
        --bg-light: #FFF8DC;
    }}
    
    /* Responsive design pour mobile */
    @media (max-width: 768px) {{
        .main .block-container {{
            padding-left: 1rem !important;
            padding-right: 1rem !important;
            padding-top: 1rem !important;
            max-width: 100% !important;
        }}
        
        /* Ajustement des cartes produits pour mobile */
        .product-card {{
            margin-bottom: 1rem !important;
            min-height: auto !important;
        }}
        
        /* Images plus petites sur mobile */
        .product-image {{
            max-width: 80px !important;
            max-height: 60px !important;
        }}
        
        /* Texte plus petit sur mobile */
        h1 {{
            font-size: 1.5rem !important;
        }}
        
        h2 {{
            font-size: 1.3rem !important;
        }}
        
        h3 {{
            font-size: 1.1rem !important;
        }}
        
        /* Boutons adaptés mobile */
        .stButton > button {{
            width: 100% !important;
            padding: 0.5rem !important;
            font-size: 0.9rem !important;
        }}
        
        /* Métriques adaptées mobile */
        [data-testid="metric-container"] {{
            padding: 0.5rem !important;
            margin-bottom: 0.5rem !important;
        }}
        
        /* Tables responsives */
        .dataframe {{
            font-size: 0.8rem !important;
        }}
        
        /* Colonnes Streamlit sur mobile */
        .row-widget {{
            flex-direction: column !important;
        }}
    }}
    
    @media (max-width: 480px) {{
        /* Très petits écrans */
        .main .block-container {{
            padding: 0.5rem !important;
        }}
        
        /* Images encore plus petites */
        .product-image {{
            max-width: 60px !important;
            max-height: 45px !important;
        }}
        
        /* Cartes produits encore plus compactes */
        .product-card {{
            padding: 0.5rem !important;
        }}
        
        h1 {{
            font-size: 1.2rem !important;
        }}
        
        .stButton > button {{
            font-size: 0.8rem !important;
            padding: 0.4rem !important;
        }}
        
        /* Sidebar plus petite */
        .css-1d391kg {{
            width: 250px !important;
        }}
    }}
    
    /* Style pour les colonnes responsives */
    .responsive-columns {{
        display: flex;
        flex-wrap: wrap;
        gap: 1rem;
    }}
    
    .responsive-column {{
        flex: 1;
        min-width: 280px;
    }}
    
    @media (max-width: 768px) {{
        .responsive-column {{
            min-width: 100%;
        }}
    }}
    
    /* Arrière-plan principal avec image pain */
    .stApp {{
        {background_image}
        min-height: 100vh;
    }}
    
    /* Arrière-plan de la sidebar */
    .css-1d391kg, .css-1y4p8pa {{
        background-color: rgba(255, 248, 220, 0.9) !important;
        backdrop-filter: blur(5px);
    }}
    
    /* Conteneur principal */
    .main .block-container {{
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(2px);
        border-radius: 1rem;
        padding: 2rem;
        margin: 1rem;
    }}
    
    /* Style des métriques */
    .metric-card {{
        background: rgba(255, 255, 255, 0.9);
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid var(--primary-color);
        box-shadow: 0 0.125rem 0.25rem rgba(0, 0, 0, 0.075);
        backdrop-filter: blur(2px);
    }}
    
    /* Style des boutons */
    .stButton > button {{
        background-color: var(--primary-color) !important;
        color: white !important;
        border-radius: 0.5rem;
        border: none;
        transition: all 0.3s ease;
        font-weight: 500;
    }}
    
    .stButton > button:hover {{
        background-color: var(--secondary-color) !important;
        transform: translateY(-1px);
        box-shadow: 0 1rem 3rem rgba(0, 0, 0, 0.175);
    }}
    
    /* Style des inputs */
    .stTextInput > div > div > input {{
        background-color: rgba(255, 255, 255, 0.9);
        border: 1px solid var(--accent-color);
        border-radius: 0.5rem;
    }}
    
    .stSelectbox > div > div > select {{
        background-color: rgba(255, 255, 255, 0.9);
        border: 1px solid var(--accent-color);
        border-radius: 0.5rem;
    }}
    
    /* Headers styling */
    h1, h2, h3 {{
        color: var(--primary-color) !important;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        text-shadow: 1px 1px 2px rgba(255, 255, 255, 0.8);
    }}
    
    /* Cards et conteneurs */
    .element-container {{
        background: rgba(255, 255, 255, 0.1);
        border-radius: 0.5rem;
        backdrop-filter: blur(2px);
    }}
    
    /* Métriques Streamlit */
    [data-testid="metric-container"] {{
        background: white !important;
        border: 1px solid var(--accent-color);
        padding: 1rem;
        border-radius: 0.5rem;
        box-shadow: 0 0.125rem 0.25rem rgba(0, 0, 0, 0.075);
    }}
    
    /* Valeurs des métriques */
    [data-testid="metric-container"] [data-testid="metric-value"] {{
        color: var(--primary-color) !important;
        font-weight: bold;
    }}
    
    /* Libellés des métriques */
    [data-testid="metric-container"] [data-testid="metric-label"] {{
        color: var(--dark-color) !important;
        font-weight: 500;
    }}
    
    /* Tables et dataframes */
    .stDataFrame {{
        background: rgba(255, 255, 255, 0.95);
        border-radius: 0.5rem;
    }}
    
    /* Alertes et messages */
    .stAlert {{
        background: rgba(255, 255, 255, 0.9);
        backdrop-filter: blur(2px);
        border-radius: 0.5rem;
    }}
    
    /* Images des produits */
    .stImage {{
        border-radius: 0.5rem;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        margin-bottom: 0.5rem;
    }}
    
    .stImage > img {{
        border-radius: 0.5rem;
        object-fit: cover;
    }}
    
    /* Formulaires */
    .stForm {{
        background: rgba(255, 255, 255, 0.9);
        padding: 1.5rem;
        border-radius: 0.5rem;
        box-shadow: 0 0.125rem 0.25rem rgba(0, 0, 0, 0.075);
        backdrop-filter: blur(2px);
    }}
    </style>
    """
    
    st.markdown(css, unsafe_allow_html=True)

# === MODÈLES DATABASE ===
def init_database():
    """Initialise la base de données SQLite avec les tables nécessaires"""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    # Table des utilisateurs
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            prenom TEXT NOT NULL,
            nom TEXT NOT NULL,
            role TEXT DEFAULT 'client',
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            entreprise_id INTEGER,
            FOREIGN KEY (entreprise_id) REFERENCES entreprises (id)
        )
    ''')
    
    # Table des entreprises
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS entreprises (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            adresse TEXT NOT NULL,
            siren TEXT UNIQUE NOT NULL,
            naf_ape TEXT NOT NULL,
            ecocert TEXT,
            numero_tva TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Table des produits
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS produits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            description TEXT,
            prix REAL NOT NULL,
            categorie TEXT DEFAULT 'pain',
            is_active BOOLEAN DEFAULT 1,
            stock INTEGER DEFAULT 0,
            stock_min INTEGER DEFAULT 5,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Table des commandes
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS commandes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT UNIQUE NOT NULL,
            client_id INTEGER NOT NULL,
            date_commande TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            date_livraison DATE NOT NULL,
            statut TEXT DEFAULT 'en_attente',
            total REAL DEFAULT 0.0,
            commentaire TEXT,
            FOREIGN KEY (client_id) REFERENCES users (id)
        )
    ''')
    
    # Table des items de commande
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS commande_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            commande_id INTEGER NOT NULL,
            produit_id INTEGER NOT NULL,
            quantite INTEGER NOT NULL,
            prix_unitaire REAL NOT NULL,
            sous_total REAL NOT NULL,
            FOREIGN KEY (commande_id) REFERENCES commandes (id),
            FOREIGN KEY (produit_id) REFERENCES produits (id)
        )
    ''')
    
    # Table des paiements
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS paiements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            commande_id INTEGER NOT NULL,
            montant REAL NOT NULL,
            mode_paiement TEXT NOT NULL,
            date_paiement DATE NOT NULL,
            reference TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (commande_id) REFERENCES commandes (id)
        )
    ''')
    
    # Table des codes promo
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS codes_promo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            type_remise TEXT NOT NULL,
            valeur REAL NOT NULL,
            date_debut DATE NOT NULL,
            date_fin DATE NOT NULL,
            usage_max INTEGER DEFAULT 100,
            usage_actuel INTEGER DEFAULT 0,
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Table des logs d'activité
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS logs_activite (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    # Table de configuration
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS configuration (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cle TEXT UNIQUE NOT NULL,
            valeur TEXT NOT NULL,
            description TEXT,
            date_modification TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Mise à jour de la table produits pour ajouter les colonnes stock si elles n'existent pas
    cursor.execute("PRAGMA table_info(produits)")
    columns = [column[1] for column in cursor.fetchall()]
    
    if 'stock' not in columns:
        cursor.execute('ALTER TABLE produits ADD COLUMN stock INTEGER DEFAULT 0')
    if 'stock_min' not in columns:
        cursor.execute('ALTER TABLE produits ADD COLUMN stock_min INTEGER DEFAULT 5')
    
    # Création des données par défaut si elles n'existent pas
    cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'admin'")
    if cursor.fetchone()[0] == 0:
        # Ajout de l'admin par défaut
        admin_password = hashlib.sha256("admin123".encode()).hexdigest()
        cursor.execute('''
            INSERT INTO users (email, password_hash, prenom, nom, role)
            VALUES (?, ?, ?, ?, ?)
        ''', ("admin@aquibiopain.com", admin_password, "Admin", "AQUI BIO PAIN", "admin"))
        
        # Ajout de produits par défaut
        produits_defaut = [
            ("Pain de campagne", "Pain traditionnel au levain", 4.50, "pain"),
            ("Baguette tradition", "Baguette française traditionnelle", 1.20, "pain"),
            ("Pain complet", "Pain complet aux graines", 5.00, "pain"),
            ("Croissant", "Viennoiserie au beurre", 1.50, "viennoiserie"),
            ("Pain au chocolat", "Viennoiserie chocolatée", 1.80, "viennoiserie")
        ]
        
        cursor.executemany('''
            INSERT INTO produits (nom, description, prix, categorie)
            VALUES (?, ?, ?, ?)
        ''', produits_defaut)
    
    conn.commit()
    conn.close()

def get_config_value(cle, valeur_defaut=""):
    """Récupère une valeur de configuration"""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT valeur FROM configuration WHERE cle = ?', (cle,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else valeur_defaut

def set_config_value(cle, valeur, description=""):
    """Sauvegarde une valeur de configuration"""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO configuration (cle, valeur, description, date_modification)
        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
    ''', (cle, valeur, description))
    conn.commit()
    conn.close()

def send_email(destinataire, sujet, contenu, smtp_config=None):
    """Envoie un email via SMTP"""
    try:
        if smtp_config is None:
            # Configuration par défaut (peut être modifiée en admin)
            smtp_config = {
                'server': get_config_value('smtp_server', 'smtp.gmail.com'),
                'port': int(get_config_value('smtp_port', '587')),
                'email': get_config_value('smtp_email', 'contact@aquibiopain.com'),
                'password': get_config_value('smtp_password', '')  # À configurer en admin
            }
        
        if not smtp_config['password']:
            return False, "Configuration SMTP incomplète"
        
        # Création du message
        msg = MIMEMultipart()
        msg['From'] = smtp_config['email']
        msg['To'] = destinataire
        msg['Subject'] = sujet
        
        # Corps du message
        msg.attach(MIMEText(contenu, 'html'))
        
        # Connexion et envoi
        server = smtplib.SMTP(smtp_config['server'], smtp_config['port'])
        server.starttls()
        server.login(smtp_config['email'], smtp_config['password'])
        server.send_message(msg)
        server.quit()
        
        return True, "Email envoyé avec succès"
        
    except Exception as e:
        return False, f"Erreur lors de l'envoi : {str(e)}"

def send_order_notification(commande_id):
    """Envoie une notification email pour une nouvelle commande"""
    if get_config_value('email_notifications') != 'True':
        return
    
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    # Récupérer les détails de la commande
    cursor.execute('''
        SELECT c.numero, c.date_commande, c.date_livraison, c.total, c.statut,
               u.prenom, u.nom, u.email
        FROM commandes c
        JOIN users u ON c.client_id = u.id
        WHERE c.id = ?
    ''', (commande_id,))
    
    commande = cursor.fetchone()
    conn.close()
    
    if commande:
        numero, date_commande, date_livraison, total, statut, prenom, nom, email = commande
        
        # Email au client
        sujet_client = f"Confirmation de commande #{numero} - SARL AQUI BIO PAIN"
        contenu_client = f"""
        <html>
        <body>
            <h2>Bonjour {prenom} {nom},</h2>
            <p>Votre commande #{numero} a été bien reçue !</p>
            <ul>
                <li><strong>Date de commande :</strong> {date_commande}</li>
                <li><strong>Date de livraison prévue :</strong> {date_livraison}</li>
                <li><strong>Montant total :</strong> {total:.2f} €</li>
                <li><strong>Statut :</strong> {statut}</li>
            </ul>
            <p>Nous vous tiendrons informé de l'évolution de votre commande.</p>
            <p>Cordialement,<br>L'équipe SARL AQUI BIO PAIN</p>
        </body>
        </html>
        """
        
        send_email(email, sujet_client, contenu_client)
        
        # Email à l'équipe
        email_admin = get_config_value('email_contact', 'contact@aquibiopain.com')
        sujet_admin = f"Nouvelle commande #{numero}"
        contenu_admin = f"""
        <html>
        <body>
            <h2>Nouvelle commande reçue</h2>
            <ul>
                <li><strong>Numéro :</strong> #{numero}</li>
                <li><strong>Client :</strong> {prenom} {nom} ({email})</li>
                <li><strong>Date de livraison :</strong> {date_livraison}</li>
                <li><strong>Montant :</strong> {total:.2f} €</li>
            </ul>
            <p>Consultez l'application pour plus de détails.</p>
        </body>
        </html>
        """
        
        send_email(email_admin, sujet_admin, contenu_admin)

def check_stock_alerts():
    """Vérifie les alertes de stock et envoie des notifications"""
    if get_config_value('stock_alerts') != 'True':
        return
    
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT nom, stock, stock_min 
        FROM produits 
        WHERE stock <= stock_min AND is_active = 1
    ''')
    
    produits_faible_stock = cursor.fetchall()
    conn.close()
    
    if produits_faible_stock:
        email_admin = get_config_value('email_contact', 'contact@aquibiopain.com')
        sujet = "Alerte Stock - SARL AQUI BIO PAIN"
        
        contenu = """
        <html>
        <body>
            <h2>⚠️ Alerte Stock</h2>
            <p>Les produits suivants ont un stock faible :</p>
            <table border="1" style="border-collapse: collapse;">
                <tr>
                    <th>Produit</th>
                    <th>Stock actuel</th>
                    <th>Stock minimum</th>
                </tr>
        """
        
        for nom, stock, stock_min in produits_faible_stock:
            contenu += f"""
                <tr>
                    <td>{nom}</td>
                    <td>{stock}</td>
                    <td>{stock_min}</td>
                </tr>
            """
        
        contenu += """
            </table>
            <p>Veuillez réapprovisionner ces produits.</p>
        </body>
        </html>
        """
        
        send_email(email_admin, sujet, contenu)

# === FONCTIONS D'AUTHENTIFICATION ===
def hash_password(password):
    """Hash un mot de passe"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    """Vérifie un mot de passe"""
    return hash_password(password) == hashed

def authenticate_user(email, password):
    """Authentifie un utilisateur"""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, email, prenom, nom, role, is_active
        FROM users 
        WHERE email = ? AND password_hash = ? AND is_active = 1
    ''', (email, hash_password(password)))
    
    user = cursor.fetchone()
    conn.close()
    
    if user:
        return {
            'id': user[0],
            'email': user[1],
            'prenom': user[2],
            'nom': user[3],
            'role': user[4],
            'is_active': user[5]
        }
    return None

# === INTERFACE DE CONNEXION ===
def show_login():
    """Affiche la page de connexion"""
    # Bouton retour à l'accueil
    if st.button("← Retour à l'accueil", key="back_home"):
        st.session_state.show_login = False
        st.rerun()
    
    st.markdown("# 🥖 SARL AQUI BIO PAIN")
    st.markdown("### Connexion à votre espace")
    
    # Message si un produit a été sélectionné
    if 'selected_product' in st.session_state:
        st.info(f"🛒 Produit sélectionné : **{st.session_state.selected_product}**  \nConnectez-vous pour continuer votre commande !")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # Onglets pour connexion et inscription
        tab1, tab2 = st.tabs(["🔑 Se connecter", "👤 Créer un compte"])
        
        with tab1:
            with st.form("login_form"):
                email = st.text_input("Adresse email", placeholder="votre@email.com")
                password = st.text_input("Mot de passe", type="password", placeholder="Votre mot de passe")
                submit_button = st.form_submit_button("Se connecter", use_container_width=True)
                
                if submit_button:
                    if email and password:
                        user = authenticate_user(email, password)
                        if user:
                            st.session_state.authenticated = True
                            st.session_state.user = user
                            st.session_state.show_login = False
                            if 'selected_product' in st.session_state:
                                del st.session_state.selected_product
                            st.success(f"Bienvenue {user['prenom']} !")
                            st.rerun()
                        else:
                            st.error("Email ou mot de passe incorrect")
                    else:
                        st.error("Veuillez remplir tous les champs")
            
            st.markdown("---")
            st.info("**Compte de démonstration:**\n\nEmail: admin@aquibiopain.com\n\nMot de passe: admin123")
        
        with tab2:
            with st.form("register_form"):
                st.markdown("#### Création de compte client")
                
                col_reg1, col_reg2 = st.columns(2)
                with col_reg1:
                    new_prenom = st.text_input("Prénom", placeholder="Votre prénom")
                    new_email = st.text_input("Email", placeholder="votre@email.com")
                with col_reg2:
                    new_nom = st.text_input("Nom", placeholder="Votre nom")
                    new_password = st.text_input("Mot de passe", type="password", placeholder="Choisissez un mot de passe")
                
                new_telephone = st.text_input("Téléphone (optionnel)", placeholder="06 12 34 56 78")
                
                # Acceptation des conditions
                accept_terms = st.checkbox("J'accepte les conditions d'utilisation et la politique de confidentialité")
                
                register_button = st.form_submit_button("Créer mon compte", use_container_width=True)
                
                if register_button:
                    if new_prenom and new_nom and new_email and new_password and accept_terms:
                        # Vérifier si l'email existe déjà
                        conn = sqlite3.connect(DATABASE_PATH)
                        cursor = conn.cursor()
                        cursor.execute("SELECT id FROM users WHERE email = ?", (new_email,))
                        
                        if cursor.fetchone():
                            st.error("❌ Un compte avec cet email existe déjà")
                        else:
                            # Créer le nouveau compte
                            password_hash = hash_password(new_password)
                            cursor.execute('''
                                INSERT INTO users (email, password_hash, prenom, nom, role, telephone)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (new_email, password_hash, new_prenom, new_nom, "client", new_telephone))
                            conn.commit()
                            
                            # Connexion automatique
                            user = authenticate_user(new_email, new_password)
                            if user:
                                st.session_state.authenticated = True
                                st.session_state.user = user
                                st.session_state.show_login = False
                                if 'selected_product' in st.session_state:
                                    del st.session_state.selected_product
                                st.success(f"✅ Compte créé avec succès ! Bienvenue {new_prenom} !")
                                st.rerun()
                        
                        conn.close()
                    else:
                        if not accept_terms:
                            st.error("❌ Veuillez accepter les conditions d'utilisation")
                        else:
                            st.error("❌ Veuillez remplir tous les champs obligatoires")

# === INTERFACE PRINCIPALE ===
def show_main_app():
    """Affiche l'application principale"""
    user = st.session_state.user
    
    # Sidebar avec navigation
    with st.sidebar:
        st.markdown(f"### Bonjour {user['prenom']} ! 👋")
        st.markdown(f"**Rôle:** {user['role'].title()}")
        
        # Navigation selon le rôle
        if user['role'] == 'admin':
            pages = {
                "🏠 Accueil": "accueil",
                "📊 Dashboard Admin": "dashboard_admin",
                "👥 Gestion Clients": "gestion_clients",
                "🏢 Gestion Entreprises": "gestion_entreprises",
                "🥖 Gestion Produits": "gestion_produits",
                "📦 Gestion Commandes": "gestion_commandes",
                "💰 Module Financier": "module_financier",
                "📊 Exports & Rapports": "exports_rapports",
                "⚙️ Administration": "administration"
            }
        else:
            pages = {
                "🏠 Accueil": "accueil",
                "👤 Mon Profil": "profil",
                "🛒 Mes Commandes": "mes_commandes",
                "📦 Nouvelle Commande": "nouvelle_commande"
            }
        
        selected_page = st.selectbox("Navigation", list(pages.keys()))
        current_page = pages[selected_page]
        
        st.markdown("---")
        if st.button("🚪 Déconnexion", use_container_width=True):
            st.session_state.authenticated = False
            st.session_state.user = None
            st.rerun()
    
    # Contenu principal selon la page sélectionnée
    if current_page == "accueil":
        show_accueil()
    elif current_page == "dashboard_admin":
        show_dashboard_admin()
    elif current_page == "gestion_clients":
        show_gestion_clients()
    elif current_page == "gestion_entreprises":
        show_gestion_entreprises()
    elif current_page == "gestion_produits":
        show_gestion_produits()
    elif current_page == "gestion_commandes":
        show_gestion_commandes()
    elif current_page == "module_financier":
        show_module_financier()
    elif current_page == "exports_rapports":
        show_exports_rapports()
    elif current_page == "administration":
        show_administration()
    elif current_page == "profil":
        show_profil()
    elif current_page == "mes_commandes":
        show_mes_commandes()
    elif current_page == "nouvelle_commande":
        show_nouvelle_commande()
    else:
        st.info(f"Page '{selected_page}' en cours de développement...")

# === PAGES DE L'APPLICATION ===
def show_accueil():
    """Page d'accueil"""
    st.markdown("# 🏠 Accueil - SARL AQUI BIO PAIN")
    
    # Affichage d'une image de présentation en haut
    try:
        pain_image_path = Path("static/images/pain1.avif")
        if pain_image_path.exists():
            st.image(str(pain_image_path), caption="Notre boulangerie artisanale", use_container_width=True)
    except Exception:
        pass
    
    # Statistiques rapides
    st.markdown("## 📊 Tableau de Bord")
    col1, col2, col3, col4 = st.columns(4)
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    with col1:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM produits WHERE is_active = 1")
        produits_count = cursor.fetchone()[0]
        st.metric("Produits actifs", produits_count, "🥖")
    
    with col2:
        cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'client'")
        clients_count = cursor.fetchone()[0]
        st.metric("Clients", clients_count, "👥")
    
    with col3:
        cursor.execute("SELECT COUNT(*) FROM commandes WHERE date_commande >= date('now', '-30 days')")
        commandes_count = cursor.fetchone()[0]
        st.metric("Commandes (30j)", commandes_count, "📦")
    
    with col4:
        cursor.execute("SELECT COALESCE(SUM(total), 0) FROM commandes WHERE date_commande >= date('now', '-30 days')")
        ca_30j = cursor.fetchone()[0]
        st.metric("CA (30j)", f"{ca_30j:.2f}€", "💰")
    
    conn.close()
    
    # Catalogue des produits
    st.markdown("## 🥖 Notre Catalogue")
    show_catalogue_produits()

def show_catalogue_produits():
    """Affiche le catalogue des produits avec images et recherche"""
    
    # Barre de recherche et filtres
    col_search, col_filter = st.columns([3, 1])
    
    with col_search:
        search_term = st.text_input("🔍 Rechercher un produit", 
                                   placeholder="Tapez le nom d'un produit...",
                                   key="search_produits")
    
    with col_filter:
        category_filter = st.selectbox("📂 Catégorie", 
                                     ["Tous", "pain", "viennoiserie", "pâtisserie", "spécialité"],
                                     key="filter_category")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Construction de la requête avec filtres
    where_conditions = ["is_active = 1"]
    params = []
    
    if search_term:
        where_conditions.append("(nom LIKE ? OR description LIKE ?)")
        params.extend([f"%{search_term}%", f"%{search_term}%"])
    
    if category_filter != "Tous":
        where_conditions.append("categorie = ?")
        params.append(category_filter)
    
    where_clause = " AND ".join(where_conditions)
    
    df_produits = pd.read_sql_query(f'''
        SELECT nom, description, prix, categorie
        FROM produits 
        WHERE {where_clause}
        ORDER BY categorie, nom
    ''', conn, params=params)
    
    conn.close()
    
    # Affichage du nombre de résultats
    if search_term or category_filter != "Tous":
        st.info(f"🔍 {len(df_produits)} produit(s) trouvé(s)")
    
    if not df_produits.empty:
        # Grouper par catégorie
        categories = df_produits['categorie'].unique()
        
        for categorie in categories:
            st.markdown(f"### {categorie.title()}")
            produits_cat = df_produits[df_produits['categorie'] == categorie]
            
            cols = st.columns(4)
            for idx, (_, produit) in enumerate(produits_cat.iterrows()):
                with cols[idx % 4]:
                    # Carte produit compacte
                    st.markdown(f"""
                    <div class="metric-card" style="padding: 0.75rem; text-align: center;">
                        <div style="margin-bottom: 0.75rem;">
                            <div style="width: 120px; height: 96px; background: #f8f9fa; border-radius: 0.5rem; display: flex; align-items: center; justify-content: center; overflow: hidden; margin: 0 auto;">
                                <img src="data:image/jpeg;base64,{get_image_base64(get_product_image(produit['nom']))}" 
                                     style="width: 100%; height: 100%; object-fit: cover; border-radius: 0.5rem;" 
                                     onerror="this.style.display='none'; this.nextElementSibling.style.display='block';">
                                <span style="display: none; color: #666; font-size: 1.5rem;">🥖</span>
                            </div>
                        </div>
                        <div>
                            <h4 style="color: var(--primary-color); margin: 0 0 0.5rem 0; font-size: 0.9rem;">{produit['nom']}</h4>
                            <p style="color: #666; margin: 0 0 0.75rem 0; font-size: 0.75rem; line-height: 1.2;">{produit['description']}</p>
                            <h3 style="color: var(--primary-color); margin: 0; font-size: 1.1rem;">{produit['prix']:.2f}€</h3>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
    else:
        st.info("Aucun produit disponible pour le moment.")

def show_dashboard_admin():
    """Dashboard administrateur"""
    st.markdown("# 📊 Dashboard Administrateur")
    
    # Graphiques et métriques détaillées
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Évolution des commandes
    st.markdown("## 📈 Évolution des Commandes")
    df_commandes = pd.read_sql_query('''
        SELECT DATE(date_commande) as date, COUNT(*) as nb_commandes, SUM(total) as ca_jour
        FROM commandes 
        WHERE date_commande >= date('now', '-30 days')
        GROUP BY DATE(date_commande)
        ORDER BY date
    ''', conn)
    
    if not df_commandes.empty:
        df_commandes['date'] = pd.to_datetime(df_commandes['date'])
        
        col1, col2 = st.columns(2)
        with col1:
            st.line_chart(df_commandes.set_index('date')['nb_commandes'])
            st.caption("Nombre de commandes par jour")
        
        with col2:
            st.line_chart(df_commandes.set_index('date')['ca_jour'])
            st.caption("Chiffre d'affaires par jour")
    
    conn.close()

def show_gestion_produits():
    """Gestion des produits avec modification et suppression"""
    st.markdown("# 🥖 Gestion des Produits")
    
    # Formulaire d'ajout de produit
    with st.expander("➕ Ajouter un nouveau produit"):
        with st.form("add_product"):
            col1, col2 = st.columns(2)
            with col1:
                nom = st.text_input("Nom du produit")
                prix = st.number_input("Prix (€)", min_value=0.01, step=0.01)
                stock = st.number_input("Stock initial", min_value=0, value=0, step=1)
            with col2:
                categorie = st.selectbox("Catégorie", ["pain", "viennoiserie", "patisserie", "sandwich"])
                description = st.text_area("Description")
                stock_min = st.number_input("Stock minimum (alerte)", min_value=0, value=5, step=1)
            
            if st.form_submit_button("Ajouter le produit"):
                if nom and prix > 0:
                    conn = sqlite3.connect(DATABASE_PATH)
                    cursor = conn.cursor()
                    cursor.execute('''
                        INSERT INTO produits (nom, description, prix, categorie, stock, stock_min)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (nom, description, prix, categorie, stock, stock_min))
                    conn.commit()
                    conn.close()
                    st.success(f"Produit '{nom}' ajouté avec succès !")
                    st.rerun()
                else:
                    st.error("Veuillez remplir tous les champs obligatoires")
    
    # Liste des produits existants avec modification
    st.markdown("## 📋 Produits existants")
    conn = sqlite3.connect(DATABASE_PATH)
    df_produits = pd.read_sql_query('''
        SELECT id, nom, description, prix, categorie, is_active, stock, stock_min
        FROM produits 
        ORDER BY categorie, nom
    ''', conn)
    
    if not df_produits.empty:
        # Sélection du produit à modifier
        col1, col2 = st.columns([2, 1])
        with col1:
            produit_names = [f"{row['nom']} - {row['categorie']} ({row['prix']:.2f}€)" for _, row in df_produits.iterrows()]
            selected_idx = st.selectbox("Sélectionner un produit à modifier", range(len(produit_names)), 
                                      format_func=lambda x: produit_names[x] if x < len(produit_names) else "")
        
        with col2:
            if st.button("🗑️ Supprimer ce produit", type="secondary"):
                if st.session_state.get('confirm_delete', False):
                    produit_id = df_produits.iloc[selected_idx]['id']
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM produits WHERE id = ?", (produit_id,))
                    conn.commit()
                    st.success("Produit supprimé !")
                    st.session_state.confirm_delete = False
                    st.rerun()
                else:
                    st.session_state.confirm_delete = True
                    st.warning("Cliquez à nouveau pour confirmer la suppression")
        
        # Formulaire de modification
        if selected_idx is not None and selected_idx < len(df_produits):
            produit = df_produits.iloc[selected_idx]
            
            with st.expander(f"✏️ Modifier '{produit['nom']}'", expanded=True):
                with st.form(f"edit_product_{produit['id']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        new_nom = st.text_input("Nom", value=produit['nom'])
                        new_prix = st.number_input("Prix (€)", value=float(produit['prix']), min_value=0.01, step=0.01)
                        new_stock = st.number_input("Stock actuel", value=int(produit.get('stock', 0)), min_value=0, step=1)
                    with col2:
                        new_categorie = st.selectbox("Catégorie", ["pain", "viennoiserie", "patisserie", "sandwich"], 
                                                   index=["pain", "viennoiserie", "patisserie", "sandwich"].index(produit['categorie']))
                        new_description = st.text_area("Description", value=produit['description'] or "")
                        new_stock_min = st.number_input("Stock minimum", value=int(produit.get('stock_min', 5)), min_value=0, step=1)
                    
                    new_actif = st.checkbox("Produit actif", value=bool(produit['is_active']))
                    
                    if st.form_submit_button("💾 Sauvegarder les modifications"):
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE produits 
                            SET nom=?, description=?, prix=?, categorie=?, is_active=?, stock=?, stock_min=?
                            WHERE id=?
                        ''', (new_nom, new_description, new_prix, new_categorie, new_actif, new_stock, new_stock_min, produit['id']))
                        conn.commit()
                        st.success("Produit modifié avec succès !")
                        st.rerun()
        
        # Tableau récapitulatif avec alertes stock
        st.markdown("## 📊 Vue d'ensemble")
        df_display = df_produits.copy()
        df_display['statut'] = df_display['is_active'].map({1: '✅ Actif', 0: '❌ Inactif'})
        df_display['alerte_stock'] = df_display.apply(
            lambda row: '🔴 Stock faible' if row.get('stock', 0) <= row.get('stock_min', 5) else '✅ OK', axis=1
        )
        df_display['stock_info'] = df_display.apply(
            lambda row: f"{row.get('stock', 0)} / {row.get('stock_min', 5)}", axis=1
        )
        
        df_final = df_display[['nom', 'description', 'prix', 'categorie', 'stock_info', 'alerte_stock', 'statut']]
        df_final.columns = ['Nom', 'Description', 'Prix (€)', 'Catégorie', 'Stock (actuel/min)', 'Alerte', 'Statut']
        
        st.dataframe(df_final, use_container_width=True)
        
        # Alertes stock faible
        stock_faible = df_produits[df_produits.get('stock', 0) <= df_produits.get('stock_min', 5)]
        if not stock_faible.empty:
            st.warning(f"⚠️ {len(stock_faible)} produit(s) en stock faible !")
            for _, prod in stock_faible.iterrows():
                st.error(f"🔴 {prod['nom']}: {prod.get('stock', 0)} unités restantes (minimum: {prod.get('stock_min', 5)})")
    
    else:
        st.info("Aucun produit trouvé.")
    
    conn.close()

def show_profil():
    """Page de profil utilisateur"""
    st.markdown("# 👤 Mon Profil")
    user = st.session_state.user
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### Informations personnelles")
        st.write(f"**Prénom:** {user['prenom']}")
        st.write(f"**Nom:** {user['nom']}")
        st.write(f"**Email:** {user['email']}")
        st.write(f"**Rôle:** {user['role'].title()}")
    
    with col2:
        st.markdown("### Actions")
        if st.button("🔧 Modifier mes informations"):
            st.session_state.edit_profile = True
            st.rerun()
        
        if st.button("🔑 Changer mon mot de passe"):
            st.session_state.change_password = True
            st.rerun()
    
    # Formulaire de modification du profil
    if st.session_state.get('edit_profile', False):
        st.markdown("---")
        st.markdown("### ✏️ Modifier mes informations")
        
        with st.form("edit_profile_form"):
            new_prenom = st.text_input("Prénom", value=user['prenom'])
            new_nom = st.text_input("Nom", value=user['nom'])
            new_email = st.text_input("Email", value=user['email'])
            
            col1, col2 = st.columns(2)
            with col1:
                if st.form_submit_button("💾 Sauvegarder", type="primary"):
                    if new_prenom and new_nom and new_email:
                        conn = sqlite3.connect(DATABASE_PATH)
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE utilisateurs 
                            SET prenom = ?, nom = ?, email = ? 
                            WHERE id = ?
                        ''', (new_prenom, new_nom, new_email, user['id']))
                        conn.commit()
                        conn.close()
                        
                        # Mettre à jour la session
                        st.session_state.user['prenom'] = new_prenom
                        st.session_state.user['nom'] = new_nom
                        st.session_state.user['email'] = new_email
                        
                        st.success("✅ Informations mises à jour avec succès !")
                        st.session_state.edit_profile = False
                        st.rerun()
                    else:
                        st.error("Tous les champs sont obligatoires")
            
            with col2:
                if st.form_submit_button("❌ Annuler"):
                    st.session_state.edit_profile = False
                    st.rerun()
    
    # Formulaire de changement de mot de passe
    if st.session_state.get('change_password', False):
        st.markdown("---")
        st.markdown("### 🔑 Changer mon mot de passe")
        
        with st.form("change_password_form"):
            current_password = st.text_input("Mot de passe actuel", type="password")
            new_password = st.text_input("Nouveau mot de passe", type="password")
            confirm_password = st.text_input("Confirmer le nouveau mot de passe", type="password")
            
            col1, col2 = st.columns(2)
            with col1:
                if st.form_submit_button("🔑 Changer le mot de passe", type="primary"):
                    if current_password and new_password and confirm_password:
                        if new_password != confirm_password:
                            st.error("Les nouveaux mots de passe ne correspondent pas")
                        elif len(new_password) < 6:
                            st.error("Le mot de passe doit contenir au moins 6 caractères")
                        else:
                            import hashlib
                            # Vérifier l'ancien mot de passe
                            hashed_current = hashlib.sha256(current_password.encode()).hexdigest()
                            conn = sqlite3.connect(DATABASE_PATH)
                            cursor = conn.cursor()
                            cursor.execute('SELECT password FROM utilisateurs WHERE id = ?', (user['id'],))
                            stored_password = cursor.fetchone()[0]
                            
                            if hashed_current == stored_password:
                                # Mettre à jour avec le nouveau mot de passe
                                hashed_new = hashlib.sha256(new_password.encode()).hexdigest()
                                cursor.execute('UPDATE utilisateurs SET password = ? WHERE id = ?', 
                                             (hashed_new, user['id']))
                                conn.commit()
                                conn.close()
                                
                                st.success("✅ Mot de passe changé avec succès !")
                                st.session_state.change_password = False
                                st.rerun()
                            else:
                                st.error("Mot de passe actuel incorrect")
                                conn.close()
                    else:
                        st.error("Tous les champs sont obligatoires")
            
            with col2:
                if st.form_submit_button("❌ Annuler"):
                    st.session_state.change_password = False
                    st.rerun()

def show_mes_commandes():
    """Affiche les commandes de l'utilisateur"""
    st.markdown("# 🛒 Mes Commandes")
    user = st.session_state.user
    
    conn = sqlite3.connect(DATABASE_PATH)
    df_commandes = pd.read_sql_query('''
        SELECT numero, date_commande, date_livraison, statut, total
        FROM commandes 
        WHERE client_id = ?
        ORDER BY date_commande DESC
    ''', conn, params=(user['id'],))
    conn.close()
    
    if not df_commandes.empty:
        st.dataframe(df_commandes, use_container_width=True)
    else:
        st.info("Vous n'avez pas encore passé de commande.")
        
        if st.button("📦 Passer ma première commande"):
            st.info("Redirection vers la page de commande...")

def show_gestion_clients():
    """Page de gestion des clients avec liaison entreprises"""
    st.markdown("# 👥 Gestion des Clients")
    
    # Formulaire d'ajout de client
    with st.expander("➕ Ajouter un nouveau client"):
        # Récupération des entreprises pour la liste déroulante
        conn = sqlite3.connect(DATABASE_PATH)
        df_entreprises = pd.read_sql_query("SELECT id, nom FROM entreprises ORDER BY nom", conn)
        
        with st.form("add_client"):
            col1, col2 = st.columns(2)
            with col1:
                prenom = st.text_input("Prénom")
                email = st.text_input("Email")
                password = st.text_input("Mot de passe temporaire", type="password")
            with col2:
                nom = st.text_input("Nom")
                if not df_entreprises.empty:
                    entreprise_options = ["Aucune"] + [f"{row['nom']}" for _, row in df_entreprises.iterrows()]
                    entreprise_idx = st.selectbox("Entreprise", range(len(entreprise_options)),
                                                format_func=lambda x: entreprise_options[x])
                    entreprise_id = df_entreprises.iloc[entreprise_idx - 1]['id'] if entreprise_idx > 0 else None
                else:
                    st.info("Aucune entreprise disponible")
                    entreprise_id = None
            
            if st.form_submit_button("Ajouter le client"):
                if prenom and nom and email and password:
                    cursor = conn.cursor()
                    
                    # Vérifier si l'email existe déjà
                    cursor.execute("SELECT id FROM users WHERE email = ?", (email,))
                    if cursor.fetchone():
                        st.error("Un utilisateur avec cet email existe déjà")
                    else:
                        cursor.execute('''
                            INSERT INTO users (email, password_hash, prenom, nom, role, entreprise_id)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (email, hash_password(password), prenom, nom, "client", entreprise_id))
                        conn.commit()
                        st.success(f"Client '{prenom} {nom}' ajouté avec succès !")
                        
                        # Log de l'action
                        user = st.session_state.user
                        cursor.execute('''
                            INSERT INTO logs_activite (user_id, action, details)
                            VALUES (?, ?, ?)
                        ''', (user['id'], "Ajout client", f"{prenom} {nom} ({email})"))
                        conn.commit()
                        st.rerun()
                else:
                    st.error("Veuillez remplir tous les champs obligatoires")
        
        conn.close()
    
    # Liste des clients existants avec modification
    st.markdown("## 📋 Liste des clients")
    conn = sqlite3.connect(DATABASE_PATH)
    df_clients = pd.read_sql_query('''
        SELECT u.id, u.prenom, u.nom, u.email, u.is_active, u.created_at,
               e.nom as entreprise, u.entreprise_id
        FROM users u
        LEFT JOIN entreprises e ON u.entreprise_id = e.id
        WHERE u.role = 'client'
        ORDER BY u.created_at DESC
    ''', conn)
    
    if not df_clients.empty:
        # Sélection d'un client à modifier
        client_options = [f"{row['prenom']} {row['nom']} ({row['email']})" for _, row in df_clients.iterrows()]
        selected_idx = st.selectbox("Sélectionner un client à modifier", 
                                  range(len(client_options)), 
                                  format_func=lambda x: client_options[x] if x < len(client_options) else "")
        
        if selected_idx is not None:
            client = df_clients.iloc[selected_idx]
            
            # Modification du client
            with st.expander(f"✏️ Modifier '{client['prenom']} {client['nom']}'"):
                df_entreprises = pd.read_sql_query("SELECT id, nom FROM entreprises ORDER BY nom", conn)
                
                with st.form(f"edit_client_{client['id']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        new_prenom = st.text_input("Prénom", value=client['prenom'])
                        new_email = st.text_input("Email", value=client['email'])
                    with col2:
                        new_nom = st.text_input("Nom", value=client['nom'])
                        if not df_entreprises.empty:
                            entreprise_options = ["Aucune"] + [f"{row['nom']}" for _, row in df_entreprises.iterrows()]
                            current_idx = 0
                            if client['entreprise_id']:
                                for i, (_, row) in enumerate(df_entreprises.iterrows()):
                                    if row['id'] == client['entreprise_id']:
                                        current_idx = i + 1
                                        break
                            
                            new_entreprise_idx = st.selectbox("Entreprise", range(len(entreprise_options)),
                                                            index=current_idx,
                                                            format_func=lambda x: entreprise_options[x],
                                                            key=f"entreprise_{client['id']}")
                            new_entreprise_id = df_entreprises.iloc[new_entreprise_idx - 1]['id'] if new_entreprise_idx > 0 else None
                        else:
                            new_entreprise_id = None
                    
                    new_actif = st.checkbox("Client actif", value=bool(client['is_active']))
                    
                    col_save, col_delete = st.columns(2)
                    with col_save:
                        if st.form_submit_button("💾 Sauvegarder"):
                            cursor = conn.cursor()
                            cursor.execute('''
                                UPDATE users 
                                SET prenom=?, nom=?, email=?, is_active=?, entreprise_id=?
                                WHERE id=?
                            ''', (new_prenom, new_nom, new_email, new_actif, new_entreprise_id, client['id']))
                            conn.commit()
                            st.success("Client modifié avec succès !")
                            st.rerun()
                    
                    with col_delete:
                        if st.form_submit_button("🗑️ Désactiver", type="secondary"):
                            cursor = conn.cursor()
                            cursor.execute("UPDATE users SET is_active = 0 WHERE id = ?", (client['id'],))
                            conn.commit()
                            st.success("Client désactivé !")
                            st.rerun()
        
        # Configuration des colonnes pour l'affichage
        df_display = df_clients.copy()
        df_display['statut'] = df_display['is_active'].map({1: '✅ Actif', 0: '❌ Inactif'})
        df_display['created_at'] = pd.to_datetime(df_display['created_at']).dt.strftime('%d/%m/%Y')
        df_display['entreprise'] = df_display['entreprise'].fillna('Particulier')
        df_display = df_display[['prenom', 'nom', 'email', 'entreprise', 'statut', 'created_at']]
        df_display.columns = ['Prénom', 'Nom', 'Email', 'Entreprise', 'Statut', 'Inscrit le']
        
        st.dataframe(df_display, use_container_width=True)
        
        # Statistiques enrichies
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total clients", len(df_clients))
        with col2:
            actifs = df_clients['is_active'].sum()
            st.metric("Clients actifs", actifs)
        with col3:
            nouveaux = len(df_clients[pd.to_datetime(df_clients['created_at']) >= pd.Timestamp.now() - pd.Timedelta(days=30)])
            st.metric("Nouveaux (30j)", nouveaux)
        with col4:
            avec_entreprise = len(df_clients[df_clients['entreprise_id'].notna()])
            st.metric("Professionnels", avec_entreprise)
    else:
        st.info("Aucun client trouvé.")
    
    conn.close()

def show_gestion_commandes():
    """Page de gestion des commandes avec workflow complet"""
    st.markdown("# 📦 Gestion des Commandes")
    
    # Onglets pour organiser les fonctionnalités
    tab1, tab2, tab3 = st.tabs(["📋 Liste des commandes", "🔄 Workflow", "📊 Statistiques"])
    
    with tab1:
        # Filtres
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            statut_filter = st.selectbox("Filtrer par statut", 
                                       ["Tous", "en_attente", "confirmee", "en_preparation", "prete", "livree", "annulee"])
        with col2:
            date_from = st.date_input("Du", value=pd.Timestamp.now().date() - pd.Timedelta(days=30))
        with col3:
            date_to = st.date_input("Au", value=pd.Timestamp.now().date())
        with col4:
            client_filter = st.text_input("Rechercher client")
        
        # Requête pour récupérer les commandes
        conn = sqlite3.connect(DATABASE_PATH)
        query = '''
            SELECT c.id, c.numero, u.prenom, u.nom, c.date_commande, 
                   c.date_livraison, c.statut, c.total, c.commentaire
            FROM commandes c
            JOIN users u ON c.client_id = u.id
            WHERE c.date_commande >= ? AND c.date_commande <= ?
        '''
        params = [str(date_from), str(date_to) + ' 23:59:59']
        
        if statut_filter != "Tous":
            query += " AND c.statut = ?"
            params.append(statut_filter)
        
        if client_filter:
            query += " AND (u.prenom LIKE ? OR u.nom LIKE ?)"
            params.extend([f"%{client_filter}%", f"%{client_filter}%"])
        
        query += " ORDER BY c.date_commande DESC"
        
        df_commandes = pd.read_sql_query(query, conn, params=params)
        
        if not df_commandes.empty:
            # Sélection d'une commande pour modification
            commande_options = [f"{row['numero']} - {row['prenom']} {row['nom']} ({row['total']:.2f}€)" 
                              for _, row in df_commandes.iterrows()]
            
            selected_cmd_idx = st.selectbox("Sélectionner une commande à modifier", 
                                          range(len(commande_options)), 
                                          format_func=lambda x: commande_options[x] if x < len(commande_options) else "")
            
            if selected_cmd_idx is not None:
                commande = df_commandes.iloc[selected_cmd_idx]
                
                # Modification du statut et autres détails
                col1, col2, col3 = st.columns(3)
                with col1:
                    statuts = ["en_attente", "confirmee", "en_preparation", "prete", "livree", "annulee"]
                    current_statut_idx = statuts.index(commande['statut']) if commande['statut'] in statuts else 0
                    new_statut = st.selectbox("Nouveau statut", statuts, index=current_statut_idx,
                                            key=f"statut_{commande['id']}")
                
                with col2:
                    new_date_livraison = st.date_input("Date de livraison", 
                                                     value=pd.to_datetime(commande['date_livraison']).date(),
                                                     key=f"date_{commande['id']}")
                
                with col3:
                    if st.button("💾 Mettre à jour", key=f"update_{commande['id']}"):
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE commandes 
                            SET statut = ?, date_livraison = ?
                            WHERE id = ?
                        ''', (new_statut, new_date_livraison, commande['id']))
                        conn.commit()
                        st.success(f"Commande {commande['numero']} mise à jour !")
                        st.rerun()
                
                # Détail de la commande sélectionnée
                with st.expander(f"📋 Détail commande {commande['numero']}"):
                    st.write(f"**Client:** {commande['prenom']} {commande['nom']}")
                    st.write(f"**Date commande:** {commande['date_commande']}")
                    st.write(f"**Date livraison:** {commande['date_livraison']}")
                    st.write(f"**Statut:** {commande['statut']}")
                    st.write(f"**Total:** {commande['total']:.2f}€")
                    if commande['commentaire']:
                        st.write(f"**Commentaire:** {commande['commentaire']}")
                    
                    # Détail des items
                    df_items = pd.read_sql_query('''
                        SELECT p.nom, ci.quantite, ci.prix_unitaire, ci.sous_total
                        FROM commande_items ci
                        JOIN produits p ON ci.produit_id = p.id
                        WHERE ci.commande_id = ?
                    ''', conn, params=(commande['id'],))
                    
                    if not df_items.empty:
                        st.markdown("**Articles commandés:**")
                        st.dataframe(df_items, use_container_width=True)
            
            # Formatage pour affichage général
            df_display = df_commandes.copy()
            df_display['date_commande'] = pd.to_datetime(df_display['date_commande']).dt.strftime('%d/%m/%Y %H:%M')
            df_display['date_livraison'] = pd.to_datetime(df_display['date_livraison']).dt.strftime('%d/%m/%Y')
            df_display['client'] = df_display['prenom'] + ' ' + df_display['nom']
            df_display['total'] = df_display['total'].apply(lambda x: f"{x:.2f}€")
            
            # Mapping des statuts avec couleurs
            statut_map = {
                'en_attente': '⏳ En attente',
                'confirmee': '✅ Confirmée',
                'en_preparation': '👨‍🍳 En préparation',
                'prete': '📦 Prête',
                'livree': '🚚 Livrée',
                'annulee': '❌ Annulée'
            }
            df_display['statut'] = df_display['statut'].map(statut_map)
            
            df_display = df_display[['numero', 'client', 'date_commande', 'date_livraison', 'statut', 'total']]
            df_display.columns = ['N° Commande', 'Client', 'Date commande', 'Date livraison', 'Statut', 'Total']
            
            st.dataframe(df_display, use_container_width=True)
        
        else:
            st.info("Aucune commande trouvée pour cette période.")
        
        conn.close()
    
    with tab2:
        st.markdown("## 🔄 Workflow des Commandes")
        
        # Tableau de bord des statuts
        conn = sqlite3.connect(DATABASE_PATH)
        df_workflow = pd.read_sql_query('''
            SELECT statut, COUNT(*) as nombre
            FROM commandes 
            WHERE date_commande >= date('now', '-7 days')
            GROUP BY statut
        ''', conn)
        
        if not df_workflow.empty:
            col1, col2, col3, col4 = st.columns(4)
            statut_colors = {
                'en_attente': ('⏳', '#FFA500'),
                'confirmee': ('✅', '#32CD32'),
                'en_preparation': ('👨‍🍳', '#1E90FF'),
                'prete': ('📦', '#9370DB'),
                'livree': ('🚚', '#228B22'),
                'annulee': ('❌', '#DC143C')
            }
            
            for i, (_, row) in enumerate(df_workflow.iterrows()):
                with [col1, col2, col3, col4][i % 4]:
                    emoji, color = statut_colors.get(row['statut'], ('📊', '#666666'))
                    st.metric(f"{emoji} {row['statut'].title()}", row['nombre'])
        
        # Actions rapides
        st.markdown("### ⚡ Actions rapides")
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("✅ Confirmer toutes les commandes en attente"):
                cursor = conn.cursor()
                cursor.execute("UPDATE commandes SET statut = 'confirmee' WHERE statut = 'en_attente'")
                affected = cursor.rowcount
                conn.commit()
                st.success(f"{affected} commande(s) confirmée(s)")
                st.rerun()
        
        with col2:
            if st.button("📦 Marquer comme prêtes les commandes en préparation"):
                cursor = conn.cursor()
                cursor.execute("UPDATE commandes SET statut = 'prete' WHERE statut = 'en_preparation'")
                affected = cursor.rowcount
                conn.commit()
                st.success(f"{affected} commande(s) marquée(s) comme prêtes")
                st.rerun()
        
        conn.close()
    
    with tab3:
        st.markdown("## 📊 Statistiques Détaillées")
        
        conn = sqlite3.connect(DATABASE_PATH)
        
        # Métriques générales
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            total_commandes = pd.read_sql_query("SELECT COUNT(*) as count FROM commandes", conn).iloc[0]['count']
            st.metric("Total commandes", total_commandes)
        
        with col2:
            ca_total = pd.read_sql_query("SELECT COALESCE(SUM(total), 0) as sum FROM commandes", conn).iloc[0]['sum']
            st.metric("CA total", f"{ca_total:.2f}€")
        
        with col3:
            panier_moyen = ca_total / total_commandes if total_commandes > 0 else 0
            st.metric("Panier moyen", f"{panier_moyen:.2f}€")
        
        with col4:
            commandes_jour = pd.read_sql_query('''
                SELECT COUNT(*) as count FROM commandes 
                WHERE DATE(date_commande) = DATE('now')
            ''', conn).iloc[0]['count']
            st.metric("Commandes aujourd'hui", commandes_jour)
        
        # Graphiques
        col1, col2 = st.columns(2)
        
        with col1:
            # Évolution des commandes sur 30 jours
            df_evolution = pd.read_sql_query('''
                SELECT DATE(date_commande) as date, COUNT(*) as nb_commandes
                FROM commandes 
                WHERE date_commande >= date('now', '-30 days')
                GROUP BY DATE(date_commande)
                ORDER BY date
            ''', conn)
            
            if not df_evolution.empty:
                df_evolution['date'] = pd.to_datetime(df_evolution['date'])
                st.line_chart(df_evolution.set_index('date')['nb_commandes'])
                st.caption("Évolution des commandes (30 derniers jours)")
        
        with col2:
            # Répartition par statut
            df_statuts = pd.read_sql_query('''
                SELECT statut, COUNT(*) as nombre
                FROM commandes 
                GROUP BY statut
            ''', conn)
            
            if not df_statuts.empty:
                st.bar_chart(df_statuts.set_index('statut')['nombre'])
                st.caption("Répartition par statut")
        
        conn.close()

def show_nouvelle_commande():
    """Page de création d'une nouvelle commande"""
    st.markdown("# 📦 Nouvelle Commande")
    user = st.session_state.user
    
    # Récupération des produits disponibles
    conn = sqlite3.connect(DATABASE_PATH)
    df_produits = pd.read_sql_query('''
        SELECT id, nom, description, prix, categorie
        FROM produits 
        WHERE is_active = 1 
        ORDER BY categorie, nom
    ''', conn)
    
    if df_produits.empty:
        st.error("Aucun produit disponible pour passer commande.")
        conn.close()
        return
    
    # Initialisation du panier en session
    if 'panier' not in st.session_state:
        st.session_state.panier = {}
    
    # Affichage des produits par catégorie
    st.markdown("## 🥖 Sélectionnez vos produits")
    
    categories = df_produits['categorie'].unique()
    
    for categorie in categories:
        st.markdown(f"### {categorie.title()}")
        produits_cat = df_produits[df_produits['categorie'] == categorie]
        
        for _, produit in produits_cat.iterrows():
            col1, col2, col3, col4, col5 = st.columns([1, 3, 2, 1, 1])
            
            # Image du produit
            with col1:
                image_path = get_product_image(produit['nom'])
                if image_path and Path(image_path).exists():
                    try:
                        st.image(image_path, width=120, caption=None)
                    except Exception:
                        st.write("📷")
                else:
                    st.write("📷")
            
            with col2:
                st.write(f"**{produit['nom']}**")
                st.write(produit['description'])
            
            with col3:
                st.write(f"**{produit['prix']:.2f}€**")
            
            with col4:
                key = f"qty_{produit['id']}"
                quantite = st.number_input("Qté", min_value=0, max_value=50, value=0, key=key)
            
            with col5:
                if quantite > 0:
                    st.session_state.panier[produit['id']] = {
                        'nom': produit['nom'],
                        'prix': produit['prix'],
                        'quantite': quantite
                    }
                elif produit['id'] in st.session_state.panier:
                    del st.session_state.panier[produit['id']]
    
    # Récapitulatif du panier
    if st.session_state.panier:
        st.markdown("## 🛒 Récapitulatif de votre commande")
        
        total = 0
        for item_id, item in st.session_state.panier.items():
            col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
            with col1:
                st.write(item['nom'])
            with col2:
                st.write(f"{item['prix']:.2f}€")
            with col3:
                st.write(f"{item['quantite']}")
            with col4:
                sous_total = item['prix'] * item['quantite']
                st.write(f"{sous_total:.2f}€")
                total += sous_total
        
        st.markdown(f"### **Total: {total:.2f}€**")
        
        # Date de livraison
        date_livraison = st.date_input("Date de livraison souhaitée", 
                                     value=pd.Timestamp.now().date() + pd.Timedelta(days=1),
                                     min_value=pd.Timestamp.now().date() + pd.Timedelta(days=1))
        
        commentaire = st.text_area("Commentaire (optionnel)")
        
        if st.button("✅ Confirmer la commande", use_container_width=True):
            # Génération du numéro de commande
            numero_commande = f"CMD{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # Insertion de la commande
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO commandes (numero, client_id, date_livraison, total, commentaire)
                VALUES (?, ?, ?, ?, ?)
            ''', (numero_commande, user['id'], date_livraison, total, commentaire))
            
            commande_id = cursor.lastrowid
            
            # Insertion des items
            for item_id, item in st.session_state.panier.items():
                cursor.execute('''
                    INSERT INTO commande_items (commande_id, produit_id, quantite, prix_unitaire, sous_total)
                    VALUES (?, ?, ?, ?, ?)
                ''', (commande_id, item_id, item['quantite'], item['prix'], item['prix'] * item['quantite']))
            
            conn.commit()
            conn.close()
            
            # Vider le panier
            st.session_state.panier = {}
            
            st.success(f"Commande {numero_commande} créée avec succès !")
            st.balloons()
    else:
        st.info("Votre panier est vide. Sélectionnez des produits ci-dessus.")
    conn.close()

def show_gestion_entreprises():
    """Page de gestion des entreprises"""
    st.markdown("# 🏢 Gestion des Entreprises")
    
    # Formulaire d'ajout d'entreprise
    with st.expander("➕ Ajouter une nouvelle entreprise"):
        with st.form("add_entreprise"):
            col1, col2 = st.columns(2)
            with col1:
                nom = st.text_input("Nom de l'entreprise")
                siren = st.text_input("Numéro SIREN (14 chiffres)")
                naf_ape = st.text_input("Code NAF/APE")
            with col2:
                adresse = st.text_area("Adresse complète")
                numero_tva = st.text_input("Numéro TVA (optionnel)")
                ecocert = st.text_input("Certification Ecocert (optionnel)")
            
            if st.form_submit_button("Ajouter l'entreprise"):
                if nom and siren and naf_ape and adresse:
                    if len(siren) == 14 and siren.isdigit():
                        conn = sqlite3.connect(DATABASE_PATH)
                        cursor = conn.cursor()
                        
                        # Vérifier si le SIREN existe déjà
                        cursor.execute("SELECT id FROM entreprises WHERE siren = ?", (siren,))
                        if cursor.fetchone():
                            st.error("Une entreprise avec ce SIREN existe déjà")
                        else:
                            cursor.execute('''
                                INSERT INTO entreprises (nom, adresse, siren, naf_ape, numero_tva, ecocert)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (nom, adresse, siren, naf_ape, numero_tva, ecocert))
                            conn.commit()
                            conn.close()
                            st.success(f"Entreprise '{nom}' ajoutée avec succès !")
                            st.rerun()
                    else:
                        st.error("Le numéro SIREN doit contenir exactement 14 chiffres")
                else:
                    st.error("Veuillez remplir tous les champs obligatoires")
    
    # Liste des entreprises
    st.markdown("## 📋 Entreprises existantes")
    conn = sqlite3.connect(DATABASE_PATH)
    df_entreprises = pd.read_sql_query('''
        SELECT e.*, COUNT(u.id) as nb_clients
        FROM entreprises e
        LEFT JOIN users u ON e.id = u.entreprise_id
        GROUP BY e.id
        ORDER BY e.nom
    ''', conn)
    
    if not df_entreprises.empty:
        # Sélection d'une entreprise à modifier
        entreprise_options = [f"{row['nom']} - SIREN: {row['siren']}" for _, row in df_entreprises.iterrows()]
        selected_idx = st.selectbox("Sélectionner une entreprise à modifier", 
                                  range(len(entreprise_options)), 
                                  format_func=lambda x: entreprise_options[x] if x < len(entreprise_options) else "")
        
        if selected_idx is not None:
            entreprise = df_entreprises.iloc[selected_idx]
            
            # Modification d'entreprise
            with st.expander(f"✏️ Modifier '{entreprise['nom']}'"):
                with st.form(f"edit_entreprise_{entreprise['id']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        new_nom = st.text_input("Nom", value=entreprise['nom'])
                        new_siren = st.text_input("SIREN", value=entreprise['siren'])
                        new_naf = st.text_input("NAF/APE", value=entreprise['naf_ape'])
                    with col2:
                        new_adresse = st.text_area("Adresse", value=entreprise['adresse'])
                        new_tva = st.text_input("Numéro TVA", value=entreprise['numero_tva'] or "")
                        new_ecocert = st.text_input("Ecocert", value=entreprise['ecocert'] or "")
                    
                    col_save, col_delete = st.columns(2)
                    with col_save:
                        if st.form_submit_button("💾 Sauvegarder"):
                            cursor = conn.cursor()
                            cursor.execute('''
                                UPDATE entreprises 
                                SET nom=?, adresse=?, siren=?, naf_ape=?, numero_tva=?, ecocert=?
                                WHERE id=?
                            ''', (new_nom, new_adresse, new_siren, new_naf, new_tva, new_ecocert, entreprise['id']))
                            conn.commit()
                            st.success("Entreprise modifiée avec succès !")
                            st.rerun()
                    
                    with col_delete:
                        if st.form_submit_button("🗑️ Supprimer", type="secondary"):
                            if entreprise['nb_clients'] == 0:
                                cursor = conn.cursor()
                                cursor.execute("DELETE FROM entreprises WHERE id = ?", (entreprise['id'],))
                                conn.commit()
                                st.success("Entreprise supprimée !")
                                st.rerun()
                            else:
                                st.error(f"Impossible de supprimer : {entreprise['nb_clients']} client(s) lié(s)")
        
        # Tableau récapitulatif
        df_display = df_entreprises[['nom', 'siren', 'naf_ape', 'numero_tva', 'ecocert', 'nb_clients']].copy()
        df_display.columns = ['Nom', 'SIREN', 'NAF/APE', 'N° TVA', 'Ecocert', 'Nb Clients']
        st.dataframe(df_display, use_container_width=True)
        
        # Statistiques
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total entreprises", len(df_entreprises))
        with col2:
            avec_clients = len(df_entreprises[df_entreprises['nb_clients'] > 0])
            st.metric("Avec clients", avec_clients)
        with col3:
            certifiees = len(df_entreprises[df_entreprises['ecocert'].notna() & (df_entreprises['ecocert'] != '')])
            st.metric("Certifiées bio", certifiees)
    
    else:
        st.info("Aucune entreprise trouvée.")
    
    conn.close()

def show_module_financier():
    """Module financier avec paiements et facturation"""
    st.markdown("# 💰 Module Financier")
    
    # Onglets pour organiser les fonctionnalités
    tab1, tab2, tab3, tab4 = st.tabs(["💳 Paiements", "📄 Factures", "📊 Statistiques", "⚙️ Paramètres"])
    
    with tab1:
        show_gestion_paiements()
    
    with tab2:
        show_gestion_factures()
    
    with tab3:
        show_statistiques_financieres()
    
    with tab4:
        show_parametres_financiers()

def show_gestion_paiements():
    """Gestion des paiements"""
    st.markdown("## � Gestion des Paiements")
    
    # Commandes non payées
    conn = sqlite3.connect(DATABASE_PATH)
    df_commandes_impayees = pd.read_sql_query('''
        SELECT c.id, c.numero, u.prenom, u.nom, c.total, c.date_commande, c.statut,
               COALESCE(SUM(p.montant), 0) as paye
        FROM commandes c
        JOIN users u ON c.client_id = u.id
        LEFT JOIN paiements p ON c.id = p.commande_id
        GROUP BY c.id
        HAVING c.total > COALESCE(SUM(p.montant), 0)
        ORDER BY c.date_commande DESC
    ''', conn)
    
    if not df_commandes_impayees.empty:
        st.markdown("### 🔴 Commandes impayées ou partiellement payées")
        
        # Sélection d'une commande pour paiement
        commande_options = [f"{row['numero']} - {row['prenom']} {row['nom']} - {row['total']:.2f}€ (payé: {row['paye']:.2f}€)" 
                           for _, row in df_commandes_impayees.iterrows()]
        
        if commande_options:
            selected_idx = st.selectbox("Commande à encaisser", range(len(commande_options)), 
                                      format_func=lambda x: commande_options[x])
            
            if selected_idx is not None:
                commande = df_commandes_impayees.iloc[selected_idx]
                reste_a_payer = commande['total'] - commande['paye']
                
                st.info(f"**Reste à payer :** {reste_a_payer:.2f}€")
                
                # Formulaire d'encaissement
                with st.form("nouveau_paiement"):
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        montant = st.number_input("Montant", min_value=0.01, max_value=float(reste_a_payer), 
                                                value=float(reste_a_payer), step=0.01)
                    
                    with col2:
                        mode_paiement = st.selectbox("Mode de paiement", 
                                                   ["especes", "carte_bancaire", "cheque", "virement", "autres"])
                    
                    with col3:
                        date_paiement = st.date_input("Date", value=datetime.now().date())
                    
                    reference = st.text_input("Référence/Numéro (optionnel)", placeholder="Numéro de chèque, transaction...")
                    commentaire = st.text_area("Commentaire (optionnel)")
                    
                    if st.form_submit_button("💰 Enregistrer le paiement"):
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO paiements (commande_id, montant, mode_paiement, date_paiement, reference, commentaire)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (commande['id'], montant, mode_paiement, date_paiement, reference, commentaire))
                        
                        # Log de l'action
                        user_id = st.session_state.user['id']
                        cursor.execute('''
                            INSERT INTO logs_activite (user_id, action, details)
                            VALUES (?, ?, ?)
                        ''', (user_id, "PAIEMENT_ENREGISTRE", f"Paiement de {montant}€ ({mode_paiement}) pour commande {commande['numero']}"))
                        
                        # Marquer la commande comme payée si total atteint
                        nouveau_total_paye = commande['paye'] + montant
                        if nouveau_total_paye >= commande['total']:
                            cursor.execute("UPDATE commandes SET statut = 'payee' WHERE id = ?", (commande['id'],))
                        
                        conn.commit()
                        st.success(f"Paiement de {montant:.2f}€ enregistré !")
                        st.rerun()
    else:
        st.success("✅ Toutes les commandes sont payées !")
    
    # Historique des paiements récents
    st.markdown("### 📋 Paiements récents")
    df_paiements = pd.read_sql_query('''
        SELECT p.*, c.numero, u.prenom, u.nom
        FROM paiements p
        JOIN commandes c ON p.commande_id = c.id
        JOIN users u ON c.client_id = u.id
        ORDER BY p.date_paiement DESC, p.created_at DESC
        LIMIT 20
    ''', conn)
    
    if not df_paiements.empty:
        df_display = df_paiements.copy()
        df_display['date_paiement'] = pd.to_datetime(df_display['date_paiement']).dt.strftime('%d/%m/%Y')
        df_display['client'] = df_display['prenom'] + ' ' + df_display['nom']
        df_display['mode_paiement'] = df_display['mode_paiement'].map({
            'especes': '💵 Espèces',
            'carte_bancaire': '💳 Carte',
            'cheque': '📄 Chèque',
            'virement': '🏦 Virement',
            'autres': '🔧 Autres'
        })
        
        df_final = df_display[['date_paiement', 'numero', 'client', 'montant', 'mode_paiement', 'reference']]
        df_final.columns = ['Date', 'N° Commande', 'Client', 'Montant (€)', 'Mode', 'Référence']
        
        st.dataframe(df_final, use_container_width=True)
    
    conn.close()

def show_gestion_factures():
    """Gestion et génération des factures"""
    st.markdown("## 📄 Gestion des Factures")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Commandes prêtes pour facturation
    df_commandes = pd.read_sql_query('''
        SELECT c.id, c.numero, u.prenom, u.nom, u.email, c.total, c.date_commande, c.statut,
               e.nom as entreprise, e.adresse, e.siren, e.numero_tva
        FROM commandes c
        JOIN users u ON c.client_id = u.id
        LEFT JOIN entreprises e ON u.entreprise_id = e.id
        WHERE c.statut IN ('confirmee', 'en_preparation', 'prete', 'livree', 'payee')
        ORDER BY c.date_commande DESC
    ''', conn)
    
    if not df_commandes.empty:
        # Sélection de commande pour facturation
        commande_options = [f"{row['numero']} - {row['prenom']} {row['nom']} - {row['total']:.2f}€" 
                           for _, row in df_commandes.iterrows()]
        
        selected_idx = st.selectbox("Commande à facturer", range(len(commande_options)), 
                                  format_func=lambda x: commande_options[x])
        
        if selected_idx is not None:
            commande = df_commandes.iloc[selected_idx]
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 📋 Détails de la commande")
                st.write(f"**N° Commande :** {commande['numero']}")
                st.write(f"**Client :** {commande['prenom']} {commande['nom']}")
                st.write(f"**Email :** {commande['email']}")
                if commande['entreprise']:
                    st.write(f"**Entreprise :** {commande['entreprise']}")
                    st.write(f"**SIREN :** {commande['siren']}")
                st.write(f"**Total :** {commande['total']:.2f}€")
            
            with col2:
                st.markdown("### 🏪 Informations boulangerie")
                # Paramètres configurables (pourraient être en base)
                nom_boulangerie = "SARL AQUI BIO PAIN"
                adresse_boulangerie = "Zone d'activité des docks maritimes\\nQuai Carriet\\n33310 LORMONT"
                siret_boulangerie = "490 057 155 RCS Bordeaux"
                tva_boulangerie = "FR 95490057155"
                
                st.write(f"**{nom_boulangerie}**")
                st.write(adresse_boulangerie.replace("\\n", "  \n"))
                st.write(f"**SIREN :** {siret_boulangerie}")
                st.write(f"**TVA :** {tva_boulangerie}")
            
            # Génération de facture
            if st.button("📄 Générer la facture PDF", use_container_width=True):
                pdf_buffer = generer_facture_pdf(commande, conn)
                
                if pdf_buffer:
                    st.success("✅ Facture générée avec succès !")
                    
                    # Bouton de téléchargement
                    st.download_button(
                        label="⬇️ Télécharger la facture",
                        data=pdf_buffer.getvalue(),
                        file_name=f"facture_{commande['numero']}.pdf",
                        mime="application/pdf"
                    )
                    
                    # Log de l'action
                    user_id = st.session_state.user['id']
                    cursor = conn.cursor()
                    cursor.execute('''
                        INSERT INTO logs_activite (user_id, action, details)
                        VALUES (?, ?, ?)
                    ''', (user_id, "FACTURE_GENEREE", f"Facture générée pour commande {commande['numero']}"))
                    conn.commit()
    
    conn.close()

def generer_facture_pdf(commande, conn):
    """Génère une facture PDF pour une commande"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Récupération des détails de la commande
        df_items = pd.read_sql_query('''
            SELECT p.nom, ci.quantite, ci.prix_unitaire, ci.sous_total
            FROM commande_items ci
            JOIN produits p ON ci.produit_id = p.id
            WHERE ci.commande_id = ?
        ''', conn, params=(commande['id'],))
        
        story = []
        styles = getSampleStyleSheet()
        
        # En-tête
        story.append(Paragraph("FACTURE", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Informations entreprise et client
        info_data = [
            ['SARL AQUI BIO PAIN', f"Facture N° : {commande['numero']}"],
            ['Zone d\'activité des docks maritimes', f"Date : {datetime.now().strftime('%d/%m/%Y')}"],
            ['Quai Carriet', ''],
            ['33310 LORMONT', ''],
            ['N° Siren : 490 057 155 RCS Bordeaux', ''],
            ['N° TVA : FR 95490057155', ''],
            ['', ''],
            ['FACTURÉ À:', ''],
            [f"{commande['prenom']} {commande['nom']}", ''],
            [commande['email'], '']
        ]
        
        if commande['entreprise']:
            info_data.append([commande['entreprise'], ''])
            if commande['adresse']:
                info_data.append([commande['adresse'], ''])
        
        info_table = Table(info_data, colWidths=[3*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 14),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 24))
        
        # Tableau des produits
        if not df_items.empty:
            # En-têtes du tableau
            data = [['Produit', 'Quantité', 'Prix unitaire', 'Total']]
            
            # Lignes de produits
            for _, item in df_items.iterrows():
                data.append([
                    item['nom'],
                    str(item['quantite']),
                    f"{item['prix_unitaire']:.2f}€",
                    f"{item['sous_total']:.2f}€"
                ])
            
            # Ligne total
            data.append(['', '', 'TOTAL:', f"{commande['total']:.2f}€"])
            
            table = Table(data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            
            story.append(table)
        
        story.append(Spacer(1, 24))
        story.append(Paragraph("Merci pour votre confiance !", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        st.error(f"Erreur lors de la génération PDF : {e}")
        return None

def show_statistiques_financieres():
    """Statistiques financières"""
    st.markdown("## 📊 Statistiques Financières")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Métriques principales
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        ca_mois = pd.read_sql_query('''
            SELECT COALESCE(SUM(total), 0) as ca
            FROM commandes 
            WHERE date_commande >= date('now', 'start of month')
        ''', conn).iloc[0]['ca']
        st.metric("CA du mois", f"{ca_mois:.2f}€")
    
    with col2:
        paiements_jour = pd.read_sql_query('''
            SELECT COALESCE(SUM(montant), 0) as total
            FROM paiements 
            WHERE date_paiement = date('now')
        ''', conn).iloc[0]['total']
        st.metric("Encaissé aujourd'hui", f"{paiements_jour:.2f}€")
    
    with col3:
        impayees = pd.read_sql_query('''
            SELECT COUNT(*) as nb
            FROM commandes c
            LEFT JOIN paiements p ON c.id = p.commande_id
            GROUP BY c.id
            HAVING c.total > COALESCE(SUM(p.montant), 0)
        ''', conn)
        nb_impayees = len(impayees) if not impayees.empty else 0
        st.metric("Commandes impayées", nb_impayees)
    
    with col4:
        panier_moyen = pd.read_sql_query('''
            SELECT AVG(total) as moyenne
            FROM commandes 
            WHERE date_commande >= date('now', '-30 days')
        ''', conn).iloc[0]['moyenne'] or 0
        st.metric("Panier moyen (30j)", f"{panier_moyen:.2f}€")
    
    # Graphiques
    st.markdown("### 📈 Évolution des ventes")
    df_ventes = pd.read_sql_query('''
        SELECT DATE(date_commande) as date, SUM(total) as ca_jour
        FROM commandes 
        WHERE date_commande >= date('now', '-30 days')
        GROUP BY DATE(date_commande)
        ORDER BY date
    ''', conn)
    
    if not df_ventes.empty:
        df_ventes['date'] = pd.to_datetime(df_ventes['date'])
        st.line_chart(df_ventes.set_index('date')['ca_jour'])
    
    conn.close()

def show_parametres_financiers():
    """Paramètres financiers"""
    st.markdown("## ⚙️ Paramètres Financiers")
    
    st.info("Configuration des informations de facturation")
    
    with st.form("config_facturation"):
        col1, col2 = st.columns(2)
        
        with col1:
            nom_boulangerie = st.text_input("Nom de la boulangerie", 
                                           value=get_config_value("nom_boulangerie", "SARL AQUI BIO PAIN"))
            adresse = st.text_area("Adresse", 
                                 value=get_config_value("adresse", "Zone d'activité des docks maritimes\\nQuai Carriet\\n33310 LORMONT"))
            siret = st.text_input("SIREN", 
                                value=get_config_value("siret", "490 057 155 RCS Bordeaux"))
        
        with col2:
            email_contact = st.text_input("Email de contact", 
                                        value=get_config_value("email_contact", "contact@aquibiopain.com"))
            telephone = st.text_input("Téléphone", 
                                    value=get_config_value("telephone", "05 56 06 92 00"))
            numero_tva = st.text_input("Numéro TVA", 
                                     value=get_config_value("numero_tva", "FR 95490057155"))
        
        if st.form_submit_button("💾 Sauvegarder"):
            # Sauvegarder en base de données
            set_config_value("nom_boulangerie", nom_boulangerie, "Nom de la boulangerie")
            set_config_value("adresse", adresse, "Adresse de la boulangerie")
            set_config_value("siret", siret, "Numéro SIREN")
            set_config_value("email_contact", email_contact, "Email de contact")
            set_config_value("telephone", telephone, "Numéro de téléphone")
            set_config_value("numero_tva", numero_tva, "Numéro de TVA")
            
            st.success("✅ Paramètres sauvegardés en base de données !")

def show_exports_rapports():
    """Module d'exports et de rapports"""
    st.markdown("# 📊 Exports et Rapports")
    
    # Onglets pour organiser les fonctionnalités
    tab1, tab2, tab3, tab4 = st.tabs(["📥 Exports Excel", "📄 Rapports PDF", "📈 Analyses", "📋 Historiques"])
    
    with tab1:
        show_exports_excel()
    
    with tab2:
        show_rapports_pdf()
    
    with tab3:
        show_analyses_donnees()
    
    with tab4:
        show_historiques()

def show_exports_excel():
    """Exports Excel"""
    st.markdown("## 📥 Exports Excel")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Section Commandes
    st.markdown("### 🛒 Export des Commandes")
    col1, col2 = st.columns(2)
    
    with col1:
        date_debut = st.date_input("Date début", value=datetime.now().date().replace(day=1))
    with col2:
        date_fin = st.date_input("Date fin", value=datetime.now().date())
    
    # Filtres avancés
    with st.expander("� Filtres avancés"):
        statuts = st.multiselect("Statuts", 
                                ["en_attente", "confirmee", "en_preparation", "prete", "livree", "payee", "annulee"],
                                default=["confirmee", "en_preparation", "prete", "livree", "payee"])
        
        clients = pd.read_sql_query("SELECT id, prenom, nom FROM users WHERE role='client'", conn)
        if not clients.empty:
            client_options = ["Tous"] + [f"{row['prenom']} {row['nom']}" for _, row in clients.iterrows()]
            client_filtre = st.selectbox("Client spécifique", client_options)
    
    if st.button("📥 Exporter les commandes", use_container_width=True):
        excel_buffer = generer_export_commandes_excel(date_debut, date_fin, statuts, conn)
        
        if excel_buffer:
            st.success("✅ Export généré avec succès !")
            st.download_button(
                label="⬇️ Télécharger commandes.xlsx",
                data=excel_buffer.getvalue(),
                file_name=f"commandes_{date_debut}_{date_fin}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    st.divider()
    
    # Section Clients
    st.markdown("### 👥 Export des Clients")
    if st.button("📥 Exporter la liste des clients", use_container_width=True):
        excel_buffer = generer_export_clients_excel(conn)
        
        if excel_buffer:
            st.success("✅ Export clients généré !")
            st.download_button(
                label="⬇️ Télécharger clients.xlsx",
                data=excel_buffer.getvalue(),
                file_name=f"clients_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    st.divider()
    
    # Section Produits
    st.markdown("### 🥖 Export des Produits")
    inclure_stocks = st.checkbox("Inclure les niveaux de stock", value=True)
    
    if st.button("📥 Exporter le catalogue produits", use_container_width=True):
        excel_buffer = generer_export_produits_excel(inclure_stocks, conn)
        
        if excel_buffer:
            st.success("✅ Export produits généré !")
            st.download_button(
                label="⬇️ Télécharger produits.xlsx",
                data=excel_buffer.getvalue(),
                file_name=f"produits_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    conn.close()

def generer_export_commandes_excel(date_debut, date_fin, statuts, conn):
    """Génère un export Excel des commandes"""
    try:
        # Requête principale
        statuts_str = "', '".join(statuts)
        query = f'''
            SELECT 
                c.numero as "N° Commande",
                c.date_commande as "Date",
                u.prenom as "Prénom",
                u.nom as "Nom",
                u.email as "Email",
                u.telephone as "Téléphone",
                e.nom as "Entreprise",
                c.statut as "Statut",
                c.total as "Total (€)",
                c.commentaires as "Commentaires"
            FROM commandes c
            JOIN users u ON c.client_id = u.id
            LEFT JOIN entreprises e ON u.entreprise_id = e.id
            WHERE c.date_commande BETWEEN '{date_debut}' AND '{date_fin}'
            AND c.statut IN ('{statuts_str}')
            ORDER BY c.date_commande DESC
        '''
        
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            st.warning("Aucune commande trouvée pour cette période.")
            return None
        
        # Formatage des données
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d/%m/%Y')
        df['Statut'] = df['Statut'].map({
            'en_attente': 'En attente',
            'confirmee': 'Confirmée',
            'en_preparation': 'En préparation',
            'prete': 'Prête',
            'livree': 'Livrée',
            'payee': 'Payée',
            'annulee': 'Annulée'
        })
        
        # Création du fichier Excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Feuille des commandes
            df.to_excel(writer, sheet_name='Commandes', index=False)
            
            # Détails des produits
            details_query = '''
                SELECT 
                    c.numero as "N° Commande",
                    p.nom as "Produit",
                    ci.quantite as "Quantité",
                    ci.prix_unitaire as "Prix unitaire (€)",
                    ci.sous_total as "Sous-total (€)"
                FROM commande_items ci
                JOIN commandes c ON ci.commande_id = c.id
                JOIN produits p ON ci.produit_id = p.id
                WHERE c.date_commande BETWEEN ? AND ?
                ORDER BY c.date_commande DESC, p.nom
            '''
            df_details = pd.read_sql_query(details_query, conn, params=(date_debut, date_fin))
            df_details.to_excel(writer, sheet_name='Détails Produits', index=False)
            
            # Mise en forme
            workbook = writer.book
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Style de l'en-tête
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Ajustement automatique des colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        st.error(f"Erreur lors de la génération de l'export : {e}")
        return None

def generer_export_clients_excel(conn):
    """Génère un export Excel des clients"""
    try:
        query = '''
            SELECT 
                u.prenom as "Prénom",
                u.nom as "Nom",
                u.email as "Email",
                u.telephone as "Téléphone",
                e.nom as "Entreprise",
                e.siren as "SIREN",
                u.date_creation as "Date d'inscription",
                COUNT(c.id) as "Nb commandes",
                COALESCE(SUM(c.total), 0) as "CA total (€)"
            FROM users u
            LEFT JOIN entreprises e ON u.entreprise_id = e.id
            LEFT JOIN commandes c ON u.id = c.client_id
            WHERE u.role = 'client'
            GROUP BY u.id
            ORDER BY u.nom, u.prenom
        '''
        
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            st.warning("Aucun client trouvé.")
            return None
        
        # Formatage
        df['Date d\'inscription'] = pd.to_datetime(df['Date d\'inscription']).dt.strftime('%d/%m/%Y')
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Clients', index=False)
            
            # Mise en forme (même style que précédemment)
            worksheet = writer.sheets['Clients']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        st.error(f"Erreur lors de la génération de l'export clients : {e}")
        return None

def generer_export_produits_excel(inclure_stocks, conn):
    """Génère un export Excel des produits"""
    try:
        if inclure_stocks:
            query = '''
                SELECT 
                    nom as "Nom",
                    description as "Description",
                    prix as "Prix (€)",
                    categorie as "Catégorie",
                    stock as "Stock actuel",
                    stock_min as "Stock minimum",
                    CASE 
                        WHEN stock <= stock_min THEN 'ALERTE'
                        WHEN stock <= stock_min * 1.5 THEN 'ATTENTION'
                        ELSE 'OK'
                    END as "État stock",
                    date_creation as "Date création"
                FROM produits
                ORDER BY categorie, nom
            '''
        else:
            query = '''
                SELECT 
                    nom as "Nom",
                    description as "Description",
                    prix as "Prix (€)",
                    categorie as "Catégorie",
                    date_creation as "Date création"
                FROM produits
                ORDER BY categorie, nom
            '''
        
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            st.warning("Aucun produit trouvé.")
            return None
        
        # Formatage
        df['Date création'] = pd.to_datetime(df['Date création']).dt.strftime('%d/%m/%Y')
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Produits', index=False)
            
            # Mise en forme avec couleurs conditionnelles pour les stocks
            worksheet = writer.sheets['Produits']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Coloration conditionnelle pour l'état des stocks
            if inclure_stocks and 'État stock' in df.columns:
                etat_col = df.columns.get_loc('État stock') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=etat_col)
                    if cell.value == 'ALERTE':
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif cell.value == 'ATTENTION':
                        cell.fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
                        cell.font = Font(bold=True)
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        st.error(f"Erreur lors de la génération de l'export produits : {e}")
        return None

def show_rapports_pdf():
    """Rapports PDF"""
    st.markdown("## 📄 Rapports PDF")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📊 Rapport mensuel")
        mois = st.selectbox("Mois", range(1, 13), value=datetime.now().month, format_func=lambda x: datetime(2023, x, 1).strftime('%B'))
        annee = st.number_input("Année", value=datetime.now().year, min_value=2020, max_value=2030)
        
        if st.button("📄 Générer rapport mensuel"):
            pdf_buffer = generer_rapport_mensuel_pdf(mois, annee, conn)
            if pdf_buffer:
                st.success("✅ Rapport généré !")
                st.download_button(
                    label="⬇️ Télécharger le rapport",
                    data=pdf_buffer.getvalue(),
                    file_name=f"rapport_mensuel_{annee}_{mois:02d}.pdf",
                    mime="application/pdf"
                )
    
    with col2:
        st.markdown("### 📈 Rapport de ventes")
        periode = st.selectbox("Période", ["7 derniers jours", "30 derniers jours", "Cette année"])
        
        if st.button("📄 Générer rapport de ventes"):
            pdf_buffer = generer_rapport_ventes_pdf(periode, conn)
            if pdf_buffer:
                st.success("✅ Rapport de ventes généré !")
                st.download_button(
                    label="⬇️ Télécharger le rapport",
                    data=pdf_buffer.getvalue(),
                    file_name=f"rapport_ventes_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf"
                )
    
    conn.close()

def generer_rapport_mensuel_pdf(mois, annee, conn):
    """Génère un rapport mensuel PDF"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Titre
        titre = f"Rapport Mensuel - {datetime(annee, mois, 1).strftime('%B %Y')}"
        story.append(Paragraph(titre, styles['Title']))
        story.append(Spacer(1, 12))
        
        # Métriques du mois
        debut_mois = f"{annee}-{mois:02d}-01"
        if mois == 12:
            fin_mois = f"{annee+1}-01-01"
        else:
            fin_mois = f"{annee}-{mois+1:02d}-01"
        
        # Statistiques principales
        stats = pd.read_sql_query(f'''
            SELECT 
                COUNT(*) as nb_commandes,
                SUM(total) as ca_total,
                AVG(total) as panier_moyen
            FROM commandes 
            WHERE date_commande >= '{debut_mois}' AND date_commande < '{fin_mois}'
        ''', conn).iloc[0]
        
        metrics_data = [
            ['Métriques', 'Valeur'],
            ['Nombre de commandes', str(stats['nb_commandes'])],
            ['Chiffre d\'affaires', f"{stats['ca_total']:.2f}€"],
            ['Panier moyen', f"{stats['panier_moyen']:.2f}€"]
        ]
        
        metrics_table = Table(metrics_data)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 12))
        
        # Top produits
        story.append(Paragraph("Top 5 des produits", styles['Heading2']))
        top_produits = pd.read_sql_query(f'''
            SELECT p.nom, SUM(ci.quantite) as total_vendu, SUM(ci.sous_total) as ca_produit
            FROM commande_items ci
            JOIN produits p ON ci.produit_id = p.id
            JOIN commandes c ON ci.commande_id = c.id
            WHERE c.date_commande >= '{debut_mois}' AND c.date_commande < '{fin_mois}'
            GROUP BY p.id
            ORDER BY total_vendu DESC
            LIMIT 5
        ''', conn)
        
        if not top_produits.empty:
            produits_data = [['Produit', 'Quantité vendue', 'CA généré']]
            for _, prod in top_produits.iterrows():
                produits_data.append([prod['nom'], str(prod['total_vendu']), f"{prod['ca_produit']:.2f}€"])
            
            produits_table = Table(produits_data)
            produits_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(produits_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        st.error(f"Erreur lors de la génération du rapport : {e}")
        return None

def generer_rapport_ventes_pdf(periode, conn):
    """Génère un rapport de ventes PDF"""
    try:
        # Déterminer la période
        if periode == "7 derniers jours":
            date_debut = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        elif periode == "30 derniers jours":
            date_debut = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        else:  # Cette année
            date_debut = f"{datetime.now().year}-01-01"
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Titre
        story.append(Paragraph(f"Rapport de Ventes - {periode}", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Ventes par jour
        ventes_jour = pd.read_sql_query(f'''
            SELECT DATE(date_commande) as jour, COUNT(*) as nb_commandes, SUM(total) as ca_jour
            FROM commandes 
            WHERE date_commande >= '{date_debut}'
            GROUP BY DATE(date_commande)
            ORDER BY jour DESC
            LIMIT 10
        ''', conn)
        
        if not ventes_jour.empty:
            story.append(Paragraph("Ventes par jour (10 derniers)", styles['Heading2']))
            ventes_data = [['Date', 'Nb Commandes', 'CA']]
            for _, vente in ventes_jour.iterrows():
                date_format = datetime.strptime(vente['jour'], '%Y-%m-%d').strftime('%d/%m/%Y')
                ventes_data.append([date_format, str(vente['nb_commandes']), f"{vente['ca_jour']:.2f}€"])
            
            ventes_table = Table(ventes_data)
            ventes_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(ventes_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        st.error(f"Erreur lors de la génération du rapport de ventes : {e}")
        return None

def show_analyses_donnees():
    """Analyses de données"""
    st.markdown("## 📈 Analyses de Données")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Graphiques et analyses
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📊 Évolution CA")
        df_ca = pd.read_sql_query('''
            SELECT DATE(date_commande) as date, SUM(total) as ca
            FROM commandes 
            WHERE date_commande >= date('now', '-30 days')
            GROUP BY DATE(date_commande)
            ORDER BY date
        ''', conn)
        
        if not df_ca.empty:
            df_ca['date'] = pd.to_datetime(df_ca['date'])
            st.line_chart(df_ca.set_index('date')['ca'])
        else:
            st.info("Pas de données disponibles")
    
    with col2:
        st.markdown("### 🥖 Top Produits")
        df_produits = pd.read_sql_query('''
            SELECT p.nom, SUM(ci.quantite) as total_vendu
            FROM commande_items ci
            JOIN produits p ON ci.produit_id = p.id
            JOIN commandes c ON ci.commande_id = c.id
            WHERE c.date_commande >= date('now', '-30 days')
            GROUP BY p.id
            ORDER BY total_vendu DESC
            LIMIT 5
        ''', conn)
        
        if not df_produits.empty:
            st.bar_chart(df_produits.set_index('nom')['total_vendu'])
        else:
            st.info("Pas de données disponibles")
    
    # Analyses détaillées
    st.markdown("### 🔍 Analyses Détaillées")
    
    tab1, tab2, tab3 = st.tabs(["🕐 Heures de pointe", "👥 Analyse clients", "💰 Rentabilité"])
    
    with tab1:
        st.info("Analyse des heures de commande (à implémenter avec horodatage)")
        
    with tab2:
        df_clients = pd.read_sql_query('''
            SELECT 
                u.prenom, u.nom,
                COUNT(c.id) as nb_commandes,
                SUM(c.total) as ca_client,
                AVG(c.total) as panier_moyen
            FROM users u
            JOIN commandes c ON u.id = c.client_id
            WHERE u.role = 'client' 
            AND c.date_commande >= date('now', '-90 days')
            GROUP BY u.id
            ORDER BY ca_client DESC
            LIMIT 10
        ''', conn)
        
        if not df_clients.empty:
            st.markdown("**Top 10 clients (90 derniers jours)**")
            df_display = df_clients.copy()
            df_display['Client'] = df_display['prenom'] + ' ' + df_display['nom']
            df_display = df_display[['Client', 'nb_commandes', 'ca_client', 'panier_moyen']]
            df_display.columns = ['Client', 'Nb Commandes', 'CA Total (€)', 'Panier Moyen (€)']
            st.dataframe(df_display, use_container_width=True)
        else:
            st.info("Pas de données clients récentes")
    
    with tab3:
        st.info("Analyse de rentabilité par produit (nécessite prix de revient)")
    
    conn.close()

def show_historiques():
    """Historiques et logs"""
    st.markdown("## 📋 Historiques")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Historique des commandes
    st.markdown("### 📦 Historique des Commandes")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        periode_commandes = st.selectbox("Période", ["7 derniers jours", "30 derniers jours", "3 derniers mois", "Tout"])
    with col2:
        statut_filtre = st.selectbox("Statut", ["Tous", "confirmee", "en_preparation", "prete", "livree", "payee", "annulee"])
    with col3:
        limite = st.number_input("Nombre max", value=50, min_value=10, max_value=1000)
    
    # Construction de la requête selon les filtres
    where_clause = "WHERE 1=1"
    if periode_commandes == "7 derniers jours":
        where_clause += " AND c.date_commande >= date('now', '-7 days')"
    elif periode_commandes == "30 derniers jours":
        where_clause += " AND c.date_commande >= date('now', '-30 days')"
    elif periode_commandes == "3 derniers mois":
        where_clause += " AND c.date_commande >= date('now', '-3 months')"
    
    if statut_filtre != "Tous":
        where_clause += f" AND c.statut = '{statut_filtre}'"
    
    df_historique = pd.read_sql_query(f'''
        SELECT 
            c.numero as "N° Commande",
            c.date_commande as "Date",
            u.prenom || ' ' || u.nom as "Client",
            c.statut as "Statut",
            c.total as "Total (€)"
        FROM commandes c
        JOIN users u ON c.client_id = u.id
        {where_clause}
        ORDER BY c.date_commande DESC
        LIMIT {limite}
    ''', conn)
    
    if not df_historique.empty:
        df_historique['Date'] = pd.to_datetime(df_historique['Date']).dt.strftime('%d/%m/%Y %H:%M')
        df_historique['Statut'] = df_historique['Statut'].map({
            'en_attente': '⏳ En attente',
            'confirmee': '✅ Confirmée',
            'en_preparation': '👨‍🍳 En préparation',
            'prete': '📦 Prête',
            'livree': '🚚 Livrée',
            'payee': '💰 Payée',
            'annulee': '❌ Annulée'
        })
        st.dataframe(df_historique, use_container_width=True)
    else:
        st.info("Aucun historique trouvé pour cette période")
    
    conn.close()

def show_administration():
    """Module d'administration système"""
    st.markdown("# ⚙️ Administration Système")
    
    # Vérification des droits admin
    if st.session_state.user['role'] != 'admin':
        st.error("🚫 Accès réservé aux administrateurs")
        return
    
    # Onglets pour organiser les fonctionnalités
    tab1, tab2, tab3, tab4 = st.tabs(["📋 Logs d'activité", "� Sauvegarde", "⚙️ Paramètres", "👥 Utilisateurs"])
    
    with tab1:
        show_logs_activite()
    
    with tab2:
        show_sauvegarde_db()
    
    with tab3:
        show_parametres_systeme()
    
    with tab4:
        show_gestion_utilisateurs()

def show_logs_activite():
    """Visualisation des logs d'activité"""
    st.markdown("## 📋 Logs d'Activité")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Filtres
    col1, col2, col3 = st.columns(3)
    
    with col1:
        date_debut = st.date_input("Date début", value=datetime.now().date() - timedelta(days=7))
    with col2:
        date_fin = st.date_input("Date fin", value=datetime.now().date())
    with col3:
        actions = pd.read_sql_query("SELECT DISTINCT action FROM logs_activite ORDER BY action", conn)['action'].tolist()
        action_filtre = st.selectbox("Action", ["Toutes"] + actions)
    
    # Utilisateurs pour filtre
    utilisateurs = pd.read_sql_query("SELECT id, prenom, nom FROM users ORDER BY nom, prenom", conn)
    if not utilisateurs.empty:
        user_options = ["Tous"] + [f"{row['prenom']} {row['nom']}" for _, row in utilisateurs.iterrows()]
        user_filtre = st.selectbox("Utilisateur", user_options)
    
    # Construction de la requête
    where_conditions = [f"DATE(timestamp) BETWEEN '{date_debut}' AND '{date_fin}'"]
    
    if action_filtre != "Toutes":
        where_conditions.append(f"action = '{action_filtre}'")
    
    if 'user_filtre' in locals() and user_filtre != "Tous":
        user_idx = user_options.index(user_filtre) - 1
        user_id = utilisateurs.iloc[user_idx]['id']
        where_conditions.append(f"l.user_id = {user_id}")
    
    where_clause = " AND ".join(where_conditions)
    
    # Récupération des logs
    df_logs = pd.read_sql_query(f'''
        SELECT 
            l.timestamp as "Date/Heure",
            u.prenom || ' ' || u.nom as "Utilisateur",
            l.action as "Action",
            l.details as "Détails"
        FROM logs_activite l
        JOIN users u ON l.user_id = u.id
        WHERE {where_clause}
        ORDER BY l.timestamp DESC
        LIMIT 500
    ''', conn)
    
    if not df_logs.empty:
        df_logs['Date/Heure'] = pd.to_datetime(df_logs['Date/Heure']).dt.strftime('%d/%m/%Y %H:%M:%S')
        
        # Coloration par type d'action
        def color_action(action):
            if 'CONNEXION' in action:
                return '🔐'
            elif 'COMMANDE' in action:
                return '🛒'
            elif 'PRODUIT' in action:
                return '🥖'
            elif 'PAIEMENT' in action:
                return '💰'
            elif 'FACTURE' in action:
                return '📄'
            else:
                return '📝'
        
        df_logs['Action'] = df_logs['Action'].apply(lambda x: f"{color_action(x)} {x}")
        
        st.dataframe(df_logs, use_container_width=True, height=400)
        
        # Statistiques des logs
        st.markdown("### 📊 Statistiques")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Total actions", len(df_logs))
        
        with col2:
            actions_uniques = df_logs['Action'].str.replace(r'^[^\s]+ ', '', regex=True).nunique()
            st.metric("Types d'actions", actions_uniques)
        
        with col3:
            users_actifs = df_logs['Utilisateur'].nunique()
            st.metric("Utilisateurs actifs", users_actifs)
        
    else:
        st.info("Aucun log trouvé pour cette période")
    
    conn.close()

def show_sauvegarde_db():
    """Sauvegarde et restauration de la base de données"""
    st.markdown("## 💾 Sauvegarde Base de Données")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📤 Sauvegarde")
        st.info("Créer une copie de sécurité de la base de données")
        
        if st.button("💾 Créer une sauvegarde", use_container_width=True):
            try:
                # Créer le répertoire de sauvegarde s'il n'existe pas
                backup_dir = Path("backups")
                backup_dir.mkdir(exist_ok=True)
                
                # Nom du fichier de sauvegarde avec timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_file = backup_dir / f"boulangerie_backup_{timestamp}.db"
                
                # Copie de la base de données
                import shutil
                shutil.copy2(DATABASE_PATH, backup_file)
                
                st.success(f"✅ Sauvegarde créée : {backup_file.name}")
                
                # Proposer le téléchargement
                with open(backup_file, 'rb') as f:
                    st.download_button(
                        label="⬇️ Télécharger la sauvegarde",
                        data=f.read(),
                        file_name=backup_file.name,
                        mime="application/octet-stream"
                    )
                
                # Log de l'action
                conn = sqlite3.connect(DATABASE_PATH)
                cursor = conn.cursor()
                user_id = st.session_state.user['id']
                cursor.execute('''
                    INSERT INTO logs_activite (user_id, action, details)
                    VALUES (?, ?, ?)
                ''', (user_id, "SAUVEGARDE_CREEE", f"Sauvegarde créée : {backup_file.name}"))
                conn.commit()
                conn.close()
                
            except Exception as e:
                st.error(f"❌ Erreur lors de la sauvegarde : {e}")
    
    with col2:
        st.markdown("### 📥 Restauration")
        st.warning("⚠️ La restauration remplacera toutes les données actuelles !")
        
        uploaded_file = st.file_uploader("Choisir un fichier de sauvegarde", type=['db'])
        
        if uploaded_file is not None:
            if st.button("🔄 Restaurer la base de données", type="secondary"):
                try:
                    # Sauvegarder l'actuelle avant restauration
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup_current = f"boulangerie_avant_restore_{timestamp}.db"
                    shutil.copy2(DATABASE_PATH, backup_current)
                    
                    # Restaurer le fichier uploadé
                    with open(DATABASE_PATH, 'wb') as f:
                        f.write(uploaded_file.getvalue())
                    
                    st.success("✅ Base de données restaurée avec succès !")
                    st.info(f"💾 Ancienne version sauvegardée : {backup_current}")
                    
                    # Redémarrage nécessaire
                    st.warning("🔄 Veuillez redémarrer l'application pour prendre en compte les changements.")
                    
                except Exception as e:
                    st.error(f"❌ Erreur lors de la restauration : {e}")
    
    # Liste des sauvegardes existantes
    st.markdown("### � Sauvegardes Existantes")
    backup_dir = Path("backups")
    if backup_dir.exists():
        backup_files = list(backup_dir.glob("*.db"))
        if backup_files:
            for backup_file in sorted(backup_files, reverse=True):
                col_name, col_size, col_date, col_action = st.columns([3, 1, 2, 1])
                
                with col_name:
                    st.text(backup_file.name)
                with col_size:
                    size_mb = backup_file.stat().st_size / (1024 * 1024)
                    st.text(f"{size_mb:.1f} MB")
                with col_date:
                    mod_time = datetime.fromtimestamp(backup_file.stat().st_mtime)
                    st.text(mod_time.strftime("%d/%m/%Y %H:%M"))
                with col_action:
                    with open(backup_file, 'rb') as f:
                        st.download_button(
                            "⬇️",
                            data=f.read(),
                            file_name=backup_file.name,
                            mime="application/octet-stream",
                            key=f"download_{backup_file.name}"
                        )
        else:
            st.info("Aucune sauvegarde trouvée")
    else:
        st.info("Répertoire de sauvegarde non créé")

def show_parametres_systeme():
    """Paramètres système"""
    st.markdown("## ⚙️ Paramètres Système")
    
    # Informations système
    st.markdown("### 📋 Informations")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.info("**Version de l'application :** 2.0.0")
        st.info("**Framework :** Streamlit")
        st.info("**Base de données :** SQLite")
    
    with col2:
        # Statistiques de la base
        conn = sqlite3.connect(DATABASE_PATH)
        
        nb_users = pd.read_sql_query("SELECT COUNT(*) as count FROM users", conn).iloc[0]['count']
        nb_produits = pd.read_sql_query("SELECT COUNT(*) as count FROM produits", conn).iloc[0]['count']
        nb_commandes = pd.read_sql_query("SELECT COUNT(*) as count FROM commandes", conn).iloc[0]['count']
        
        st.metric("Utilisateurs", nb_users)
        st.metric("Produits", nb_produits)
        st.metric("Commandes", nb_commandes)
        
        conn.close()
    
    st.divider()
    
    # Paramètres configurables
    st.markdown("### 🔧 Configuration")
    
    with st.form("parametres_systeme"):
        col1, col2 = st.columns(2)
        
        with col1:
            nom_boulangerie = st.text_input("Nom de la boulangerie", 
                                           value=get_config_value("nom_boulangerie", "SARL AQUI BIO PAIN"))
            adresse_boulangerie = st.text_area("Adresse", 
                                             value=get_config_value("adresse_boulangerie", "Zone d'activité des docks maritimes\\nQuai Carriet\\n33310 LORMONT"))
            email_contact = st.text_input("Email de contact", 
                                        value=get_config_value("email_contact", "contact@aquibiopain.com"))
        
        with col2:
            telephone = st.text_input("Téléphone", 
                                    value=get_config_value("telephone", "05 56 06 92 00"))
            siret = st.text_input("SIREN", 
                                value=get_config_value("siret", "490 057 155 RCS Bordeaux"))
            numero_tva = st.text_input("Numéro TVA", 
                                     value=get_config_value("numero_tva", "FR 95490057155"))
        
        # Options système
        st.subheader("Options système")
        col3, col4 = st.columns(2)
        
        with col3:
            stock_alerts = st.checkbox("Alertes de stock", value=get_config_value("stock_alerts") == "True")
            auto_backup = st.checkbox("Sauvegarde automatique", value=get_config_value("auto_backup") == "True")
        
        with col4:
            email_notifications = st.checkbox("Notifications email", value=get_config_value("email_notifications") == "True")
            maintenance_mode = st.checkbox("Mode maintenance", value=get_config_value("maintenance_mode") == "True")
        
        if st.form_submit_button("💾 Sauvegarder les paramètres"):
            # Sauvegarder en base de données
            set_config_value("nom_boulangerie", nom_boulangerie, "Nom de la boulangerie")
            set_config_value("adresse_boulangerie", adresse_boulangerie, "Adresse de la boulangerie")
            set_config_value("email_contact", email_contact, "Email de contact")
            set_config_value("telephone", telephone, "Numéro de téléphone")
            set_config_value("siret", siret, "Numéro SIREN")
            set_config_value("numero_tva", numero_tva, "Numéro de TVA")
            set_config_value("stock_alerts", str(stock_alerts), "Alertes de stock")
            set_config_value("auto_backup", str(auto_backup), "Sauvegarde automatique")
            set_config_value("email_notifications", str(email_notifications), "Notifications email")
            set_config_value("maintenance_mode", str(maintenance_mode), "Mode maintenance")
            
            st.success("✅ Paramètres sauvegardés en base de données !")
    
    # Configuration SMTP pour les emails
    st.markdown("### 📧 Configuration Email")
    with st.form("config_smtp"):
        st.info("Configuration pour l'envoi automatique d'emails")
        
        col1, col2 = st.columns(2)
        with col1:
            smtp_server = st.text_input("Serveur SMTP", 
                                       value=get_config_value("smtp_server", "smtp.gmail.com"))
            smtp_port = st.number_input("Port SMTP", 
                                       value=int(get_config_value("smtp_port", "587")),
                                       min_value=1, max_value=65535)
        
        with col2:
            smtp_email = st.text_input("Email d'envoi", 
                                      value=get_config_value("smtp_email", "contact@aquibiopain.com"))
            smtp_password = st.text_input("Mot de passe email", 
                                         type="password",
                                         value=get_config_value("smtp_password", ""))
        
        if st.form_submit_button("💾 Sauvegarder config SMTP"):
            set_config_value("smtp_server", smtp_server, "Serveur SMTP")
            set_config_value("smtp_port", str(smtp_port), "Port SMTP")
            set_config_value("smtp_email", smtp_email, "Email d'envoi")
            if smtp_password:
                set_config_value("smtp_password", smtp_password, "Mot de passe SMTP")
            
            st.success("✅ Configuration SMTP sauvegardée !")
            
        # Test d'envoi
        if st.form_submit_button("📧 Tester l'envoi d'email"):
            if smtp_email and smtp_password:
                success, message = send_email(
                    smtp_email, 
                    "Test - SARL AQUI BIO PAIN", 
                    "<p>Ceci est un email de test pour vérifier la configuration SMTP.</p>"
                )
                if success:
                    st.success("✅ Email de test envoyé avec succès !")
                else:
                    st.error(f"❌ Erreur : {message}")
            else:
                st.warning("Veuillez configurer l'email et le mot de passe SMTP")

def show_gestion_utilisateurs():
    """Gestion des utilisateurs"""
    st.markdown("## 👥 Gestion des Utilisateurs")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Liste des utilisateurs
    df_users = pd.read_sql_query('''
        SELECT 
            id,
            prenom as "Prénom",
            nom as "Nom", 
            email as "Email",
            telephone as "Téléphone",
            role as "Rôle",
            date_creation as "Date création"
        FROM users
        ORDER BY role, nom, prenom
    ''', conn)
    
    if not df_users.empty:
        # Formatage
        df_display = df_users.copy()
        df_display['Date création'] = pd.to_datetime(df_display['Date création']).dt.strftime('%d/%m/%Y')
        df_display['Rôle'] = df_display['Rôle'].map({
            'admin': '👑 Admin',
            'client': '👤 Client'
        })
        
        # Affichage avec possibilité de sélection
        selected_rows = st.dataframe(
            df_display.drop('id', axis=1),
            use_container_width=True,
            selection_mode="single-row",
            on_select="rerun"
        )
        
        # Actions sur utilisateur sélectionné
        if hasattr(selected_rows, 'selection') and selected_rows.selection.rows:
            selected_idx = selected_rows.selection.rows[0]
            selected_user = df_users.iloc[selected_idx]
            
            st.markdown(f"### Actions pour {selected_user['Prénom']} {selected_user['Nom']}")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("✏️ Modifier", use_container_width=True):
                    st.session_state.edit_user_id = selected_user['id']
            
            with col2:
                if selected_user['Rôle'] != 'admin':  # Ne pas supprimer les admins
                    if st.button("🗑️ Supprimer", use_container_width=True, type="secondary"):
                        if st.session_state.get('confirm_delete') == selected_user['id']:
                            # Suppression confirmée
                            cursor = conn.cursor()
                            cursor.execute("DELETE FROM users WHERE id = ?", (selected_user['id'],))
                            conn.commit()
                            st.success("✅ Utilisateur supprimé")
                            st.rerun()
                        else:
                            st.session_state.confirm_delete = selected_user['id']
                            st.warning("⚠️ Cliquez à nouveau pour confirmer la suppression")
            
            with col3:
                if selected_user['Rôle'] == '👤 Client':
                    if st.button("👑 Promouvoir Admin", use_container_width=True):
                        cursor = conn.cursor()
                        cursor.execute("UPDATE users SET role = 'admin' WHERE id = ?", (selected_user['id'],))
                        conn.commit()
                        st.success("✅ Utilisateur promu administrateur")
                        st.rerun()
    
    else:
        st.info("Aucun utilisateur trouvé")
    
    st.divider()
    
    # Ajout d'un nouvel utilisateur
    st.markdown("### ➕ Ajouter un Utilisateur")
    
    with st.form("nouveau_utilisateur"):
        col1, col2 = st.columns(2)
        
        with col1:
            new_prenom = st.text_input("Prénom")
            new_nom = st.text_input("Nom")
            new_email = st.text_input("Email")
        
        with col2:
            new_telephone = st.text_input("Téléphone")
            new_role = st.selectbox("Rôle", ["client", "admin"])
            new_password = st.text_input("Mot de passe", type="password")
        
        if st.form_submit_button("👤 Créer l'utilisateur"):
            if all([new_prenom, new_nom, new_email, new_password]):
                try:
                    cursor = conn.cursor()
                    password_hash = hashlib.sha256(new_password.encode()).hexdigest()
                    
                    cursor.execute('''
                        INSERT INTO users (prenom, nom, email, telephone, password, role)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (new_prenom, new_nom, new_email, new_telephone, password_hash, new_role))
                    
                    conn.commit()
                    st.success(f"✅ Utilisateur {new_prenom} {new_nom} créé avec succès !")
                    st.rerun()
                    
                except sqlite3.IntegrityError:
                    st.error("❌ Cet email est déjà utilisé")
                except Exception as e:
                    st.error(f"❌ Erreur lors de la création : {e}")
            else:
                st.error("❌ Veuillez remplir tous les champs obligatoires")
    
    conn.close()

# Ajout des nouvelles tables pour les fonctionnalités avancées
def show_public_homepage():
    """Page d'accueil publique accessible sans connexion"""
    # En-tête avec navigation
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        st.markdown("# 🥖 SARL AQUI BIO PAIN")
        st.markdown("*Votre boulangerie artisanale bio*")
    
    with col2:
        if st.button("🛒 Commander", use_container_width=True):
            st.session_state.show_login = True
            st.rerun()
    
    with col3:
        if st.button("🔑 Se connecter", use_container_width=True):
            st.session_state.show_login = True
            st.rerun()
    
    # Image de présentation
    try:
        pain_image_path = Path("static/images/pain1.avif")
        if pain_image_path.exists():
            st.image(str(pain_image_path), caption="Notre boulangerie artisanale", use_container_width=True)
    except Exception:
        pass
    
    # Présentation de la boulangerie
    st.markdown("## 🌾 Notre Histoire")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **SARL AQUI BIO PAIN** est votre boulangerie artisanale bio située à Lormont. 
        Nous vous proposons des produits frais, fabriqués avec des ingrédients biologiques 
        et selon les méthodes traditionnelles.
        
        **Nos valeurs :**
        - 🌱 Produits 100% biologiques
        - 👨‍🍳 Fabrication artisanale
        - 🕐 Fraîcheur garantie
        - 🎯 Qualité premium
        """)
    
    with col2:
        st.markdown("""
        **📍 Notre adresse :**  
        Zone d'activité des docks maritimes  
        Quai Carriet  
        33310 LORMONT
        
        **📞 Contact :**  
        Tél : 05 56 06 92 00  
        Email : contact@aquibiopain.com
        
        **🕰️ Horaires :**  
        Lun-Sam : 6h30 - 19h30  
        Dimanche : 7h00 - 13h00
        """)

def get_image_base64(image_path):
    """Convertit une image en base64 pour l'affichage HTML"""
    if not image_path or not Path(image_path).exists():
        return ""
    
    try:
        import base64
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    except Exception:
        return ""
    
def get_product_image(nom_produit):
    """Retourne le chemin de l'image correspondant au produit"""
    # Mapping des noms de produits vers les images
    product_images = {
        # Pain complet
        'pain complet': 'static/images/complet.jpg',
        'complet': 'static/images/complet.jpg',
        
        # Pain de campagne
        'pain de campagne': 'static/images/compagne.webp',
        'campagne': 'static/images/compagne.webp',
        'tourte': 'static/images/compagne.webp',
        
        # Pain traditionnel/baguette tradition
        'pain traditionnel': 'static/images/tradi.jpg',
        'baguette tradition': 'static/images/tradi.jpg',
        'tradition': 'static/images/tradi.jpg',
        'tradi': 'static/images/tradi.jpg',
        
        # Pain au chocolat
        'pain au chocolat': 'static/images/choco.jpg',
        'chocolat': 'static/images/choco.jpg',
        'choco': 'static/images/choco.jpg',
        
        # Croissant/viennoiserie
        'croissant': 'static/images/croi.webp',
        'croi': 'static/images/croi.webp',
        'viennoiserie': 'static/images/croi.webp',
        
        # Images par défaut pour autres produits
        'pain': 'static/images/pain1.avif',
        'baguette': 'static/images/tradi.jpg'  # Baguettes → image tradition par défaut
    }
    
    # Recherche par correspondance exacte ou partielle
    nom_lower = nom_produit.lower()
    
    # Correspondance exacte d'abord
    if nom_lower in product_images:
        return product_images[nom_lower]
    
    # Correspondance partielle ensuite
    for keyword, image_path in product_images.items():
        if keyword in nom_lower:
            image_path_obj = Path(image_path)
            if image_path_obj.exists():
                return image_path
    
    # Image par défaut
    default_image = Path('static/images/pain1.avif')
    if default_image.exists():
        return str(default_image)
    
    return None

def show_public_homepage():
    """Page d'accueil publique accessible sans connexion"""
    # En-tête avec navigation
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        st.markdown("# 🥖 SARL AQUI BIO PAIN")
        st.markdown("*Votre boulangerie artisanale bio*")
    
    with col2:
        if st.button("🛒 Commander", use_container_width=True):
            st.session_state.show_login = True
            st.rerun()
    
    with col3:
        if st.button("🔑 Se connecter", use_container_width=True):
            st.session_state.show_login = True
            st.rerun()
    
    # Image de présentation
    try:
        pain_image_path = Path("static/images/pain1.avif")
        if pain_image_path.exists():
            st.image(str(pain_image_path), caption="Notre boulangerie artisanale", use_container_width=True)
    except Exception:
        pass
    
    # Présentation de la boulangerie
    st.markdown("## 🌾 Notre Histoire")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **SARL AQUI BIO PAIN** est votre boulangerie artisanale bio située à Lormont. 
        Nous vous proposons des produits frais, fabriqués avec des ingrédients biologiques 
        et selon les méthodes traditionnelles.
        
        **Nos valeurs :**
        - 🌱 Produits 100% biologiques
        - 👨‍🍳 Fabrication artisanale
        - 🕐 Fraîcheur garantie
        - 🎯 Qualité premium
        """)
    
    with col2:
        st.markdown("""
        **📍 Notre adresse :**  
        Zone d'activité des docks maritimes  
        Quai Carriet  
        33310 LORMONT
        
        **📞 Contact :**  
        Tél : 05 56 06 92 00  
        Email : contact@aquibiopain.com
        
        **🕰️ Horaires :**  
        Lun-Sam : 6h30 - 19h30  
        Dimanche : 7h00 - 13h00
        """)
    
    # Catalogue des produits PUBLIQUE avec images et recherche
    st.markdown("## 🥖 Notre Catalogue")
    
    # Barre de recherche et filtres pour la page publique
    col_search, col_filter = st.columns([3, 1])
    
    with col_search:
        search_term_public = st.text_input("🔍 Rechercher un produit", 
                                          placeholder="Tapez le nom d'un produit...",
                                          key="search_produits_public")
    
    with col_filter:
        category_filter_public = st.selectbox("📂 Catégorie", 
                                            ["Tous", "pain", "viennoiserie", "pâtisserie", "spécialité"],
                                            key="filter_category_public")
    
    conn = sqlite3.connect(DATABASE_PATH)
    
    # Construction de la requête avec filtres pour la page publique
    where_conditions = ["is_active = 1"]
    params = []
    
    if search_term_public:
        where_conditions.append("(nom LIKE ? OR description LIKE ?)")
        params.extend([f"%{search_term_public}%", f"%{search_term_public}%"])
    
    if category_filter_public != "Tous":
        where_conditions.append("categorie = ?")
        params.append(category_filter_public)
    
    where_clause = " AND ".join(where_conditions)
    
    df_produits = pd.read_sql_query(f'''
        SELECT nom, description, prix, categorie
        FROM produits 
        WHERE {where_clause}
        ORDER BY categorie, nom
    ''', conn, params=params)
    
    conn.close()
    
    # Affichage du nombre de résultats
    if search_term_public or category_filter_public != "Tous":
        st.info(f"🔍 {len(df_produits)} produit(s) trouvé(s)")
    
    if not df_produits.empty:
        # Grouper par catégorie
        categories = df_produits['categorie'].unique()
        
        for categorie in categories:
            st.markdown(f"### {categorie.title()}")
            produits_cat = df_produits[df_produits['categorie'] == categorie]
            
            # Affichage en grille avec cartes plus compactes (4 colonnes)
            cols = st.columns(4)
            for idx, (_, produit) in enumerate(produits_cat.iterrows()):
                with cols[idx % 4]:
                    with st.container():
                        # Carte produit avec image et infos côte à côte
                        st.markdown(f"""
                        <div class="metric-card" style="padding: 0.75rem; text-align: center;">
                            <div style="margin-bottom: 0.75rem;">
                                <div style="width: 120px; height: 96px; background: #f8f9fa; border-radius: 0.5rem; display: flex; align-items: center; justify-content: center; overflow: hidden; margin: 0 auto;">
                                    <img src="data:image/jpeg;base64,{get_image_base64(get_product_image(produit['nom']))}" 
                                         style="width: 100%; height: 100%; object-fit: cover; border-radius: 0.5rem;" 
                                         onerror="this.style.display='none'; this.nextElementSibling.style.display='block';">
                                    <span style="display: none; color: #666; font-size: 1.5rem;">🥖</span>
                                </div>
                            </div>
                            <div>
                                <h4 style="color: var(--primary-color); margin: 0 0 0.5rem 0; font-size: 0.9rem;">{produit['nom']}</h4>
                                <p style="color: #666; margin: 0 0 0.75rem 0; font-size: 0.75rem; line-height: 1.2;">{produit['description']}</p>
                                <h3 style="color: var(--primary-color); margin: 0; font-size: 1.1rem;">{produit['prix']:.2f}€</h3>
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if st.button(f"🛒 Commander", key=f"cmd_{produit['nom']}", use_container_width=True):
                            st.session_state.show_login = True
                            st.session_state.selected_product = produit['nom']
                            st.info(f"💡 Connectez-vous pour commander '{produit['nom']}'")
                            st.rerun()
    else:
        st.info("Notre catalogue est en cours de mise à jour.")
    
    # Call-to-action
    st.markdown("---")
    st.markdown("## 🚀 Prêt à commander ?")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🛒 COMMENCER MA COMMANDE", use_container_width=True, type="primary"):
            st.session_state.show_login = True
            st.rerun()
        
        st.markdown("<center><small>Créez votre compte en quelques secondes</small></center>", unsafe_allow_html=True)
    
    # Informations légales en footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 0.8rem;'>
        SARL AQUI BIO PAIN - SIREN: 490 057 155 RCS Bordeaux - TVA: FR 95490057155<br>
        Zone d'activité des docks maritimes, Quai Carriet, 33310 LORMONT
    </div>
    """, unsafe_allow_html=True)

def main():
    """Fonction principale de l'application"""
    # Initialisation de la base de données
    init_database()
    
    # Chargement du CSS
    load_css()
    
    # Initialisation des variables de session
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    if 'user' not in st.session_state:
        st.session_state.user = None
    if 'show_login' not in st.session_state:
        st.session_state.show_login = False
    
    # Logique d'affichage
    if st.session_state.authenticated:
        show_main_app()
    elif st.session_state.show_login:
        show_login()
    else:
        show_public_homepage()

if __name__ == "__main__":
    main()